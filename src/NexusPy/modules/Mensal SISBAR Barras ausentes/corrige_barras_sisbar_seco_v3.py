# -*- coding: utf-8 -*-
"""
Automação SISBAR SECO - Correção de Barras Ausentes
Versão: 3.0.0
Data: 03/02/2026

Sistema completo com:
- Abas de Logs, Preview de Abas, Checklist e Tabela Final
- Visualização em tempo real do processamento
- Interface moderna com abas organizadas

Requisitos:
    pip install pandas openpyxl PySide6
"""

from __future__ import annotations
import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# GUI
try:
    from PySide6.QtWidgets import (
        QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
        QPushButton, QFileDialog, QTextEdit, QComboBox, QMessageBox, 
        QCheckBox, QGroupBox, QSpinBox, QTabWidget, QTableWidget, 
        QTableWidgetItem, QHeaderView, QProgressBar, QSplitter
    )
    from PySide6.QtCore import Qt, QThread, Signal
    from PySide6.QtGui import QFont, QColor
except ImportError:
    print("ERRO: PySide6 não instalado. Execute: pip install PySide6")
    sys.exit(1)


# ================================ CONSTANTES ================================ #

PT_MONTHS = [
    "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
    "Jul", "Ago", "Set", "Out", "Nov", "Dez"
]

DEFAULT_ABAS = [
    "Futuras Ligadas",
    "Futuras Desligadas", 
    "Desativadas",
    "Ativas Desligadas 01",
    "Ativas Desligadas 02",
    "Ativas Faltantes 01",
    "Ativas Faltantes 02"
]

AREAS_SECO_DEFAULT = [10, 11, 12, 20, 21, 22, 30, 31, 32]


# ================================ DATACLASSES =============================== #

@dataclass
class ConfigFormulas:
    """Configurações para geração de fórmulas Excel"""
    sep: str
    parecer_file: str
    parecer_sheet_nome: str
    parecer_range: str
    mapping_sheet: str
    mapping_range: str
    fallback_aba: str
    fallback_text: str


@dataclass
class DadosMensalAnterior:
    """Dados extraídos do mensal anterior"""
    df: pd.DataFrame
    arquivo_origem: str
    mes_ano_origem: Tuple[str, str]


# ================================ UTILITÁRIOS =============================== #

class Utils:
    """Funções utilitárias"""
    
    @staticmethod
    def extrair_mes_ano(filename: str) -> Optional[Tuple[str, str]]:
        """Extrai mês e ano de um nome de arquivo"""
        pattern = r"(?i)mensal\s*[_-]?\s*([A-Za-z]{3})(\d{2})\.xlsx$"
        match = re.search(pattern, os.path.basename(filename))
        if match:
            return (match.group(1).capitalize(), match.group(2))
        return None
    
    @staticmethod
    def calcular_proximo_mes(mes: str, ano: str) -> Tuple[str, str]:
        """Calcula o próximo mês/ano"""
        try:
            idx = PT_MONTHS.index(mes)
            proximo_idx = (idx + 1) % 12
            proximo_mes = PT_MONTHS[proximo_idx]
            proximo_ano = str(int(ano) + 1) if proximo_idx == 0 else ano
            return (proximo_mes, proximo_ano)
        except ValueError:
            raise ValueError(f"Mês inválido: {mes}")
    
    @staticmethod
    def gerar_nome_proximo_mensal(arquivo_anterior: str) -> Optional[str]:
        """Gera o nome do próximo arquivo mensal"""
        mes_ano = Utils.extrair_mes_ano(arquivo_anterior)
        if not mes_ano:
            return None
        
        mes, ano = mes_ano
        proximo_mes, proximo_ano = Utils.calcular_proximo_mes(mes, ano)
        return f"Mensal - {proximo_mes}{proximo_ano}.xlsx"
    
    @staticmethod
    def excel_quote_external_ref(caminho: str, aba: str) -> str:
        """Cria referência externa para fórmula Excel"""
        norm = os.path.normpath(caminho)
        base = os.path.basename(norm)
        folder = os.path.dirname(norm)
        ref = f"{folder}\\[{base}]{aba}"
        ref = ref.replace("'", "''")
        return f"'{ref}'!"


# ============================== PROCESSADOR ================================= #

class ProcessadorSISBAR:
    """Classe principal para processar os dados SISBAR"""
    
    def __init__(self, logger_fn=print, progress_fn=None):
        self.log = logger_fn
        self.progress = progress_fn or (lambda x: None)
    
    def carregar_mensal_anterior(self, caminho: str) -> DadosMensalAnterior:
        """Carrega dados do mensal anterior"""
        self.log(f"📖 Lendo: {os.path.basename(caminho)}")
        self.progress(10)
        
        df_raw = pd.read_excel(caminho, sheet_name=0, engine='openpyxl')
        mapeamento = self._detectar_colunas(df_raw.columns)
        
        df = pd.DataFrame()
        for col_padrao, col_encontrada in mapeamento.items():
            if col_encontrada:
                df[col_padrao] = df_raw[col_encontrada[0]]
            else:
                df[col_padrao] = ""
        
        n_antes = len(df)
        df = df[~df['Nº Barra'].isna() & (df['Nº Barra'] != "")]
        df = df.reset_index(drop=True)
        
        self.log(f"✓ Linhas carregadas: {len(df)} (de {n_antes})")
        self.progress(20)
        
        mes_ano = Utils.extrair_mes_ano(caminho)
        if not mes_ano:
            mes_ano = ("", "")
        
        return DadosMensalAnterior(
            df=df,
            arquivo_origem=caminho,
            mes_ano_origem=mes_ano
        )
    
    def _detectar_colunas(self, colunas: List[str]) -> Dict[str, List[str]]:
        """Detecta colunas do Excel"""
        lower_map = {c.lower(): c for c in colunas}
        
        def encontrar(*opcoes: str) -> List[str]:
            for opcao in opcoes:
                if opcao.lower() in lower_map:
                    return [lower_map[opcao.lower()]]
            for col in colunas:
                if any(opt in col.lower() for opt in opcoes):
                    return [col]
            return []
        
        return {
            'Nº Barra': encontrar('nº barra', 'no barra', 'num barra', 'id barra', 'nº'),
            'Nome': encontrar('nome', 'nome barra'),
            'Agente': encontrar('agente', 'empresa', 'concessionária'),
            'Área': encontrar('área', 'area', 'região', 'regiao'),
        }
    
    def ler_valores_a1(self, arquivo: str, abas: List[str]) -> Dict[str, str]:
        """Lê valores da célula A1 de cada aba"""
        self.log(f"📋 Lendo A1 de {len(abas)} aba(s)...")
        self.progress(30)
        resultado = {}
        
        try:
            wb = load_workbook(arquivo, data_only=True)
            for aba in abas:
                if aba in wb.sheetnames:
                    valor = wb[aba]['A1'].value
                    resultado[aba] = str(valor) if valor else ""
                    self.log(f"  • {aba}: '{resultado[aba][:50]}...'")
                else:
                    resultado[aba] = ""
                    self.log(f"  ⚠ {aba}: não encontrada")
            wb.close()
            self.progress(40)
        except Exception as e:
            self.log(f"❌ Erro ao ler A1: {e}")
            return {}
        
        return resultado
    
    def carregar_mapeamento_area(self, arquivo: str, aba: str) -> Optional[pd.DataFrame]:
        """Carrega mapeamento Nº Barra -> Área"""
        try:
            df = pd.read_excel(arquivo, sheet_name=aba, engine='openpyxl')
            
            col_barra = None
            col_area = None
            
            for col in df.columns:
                col_lower = str(col).lower()
                if 'barra' in col_lower and 'n' in col_lower:
                    col_barra = col
                elif 'área' in col_lower or 'area' in col_lower:
                    col_area = col
            
            if not col_barra or not col_area:
                self.log("⚠ Colunas de mapeamento não encontradas")
                return None
            
            df_map = df[[col_barra, col_area]].rename(
                columns={col_barra: 'Nº Barra', col_area: 'Área'}
            )
            
            self.log(f"✓ Mapeamento carregado: {len(df_map)} linhas")
            return df_map
            
        except Exception as e:
            self.log(f"❌ Erro ao carregar mapeamento: {e}")
            return None
    
    def gerar_mensal(self,
                     dados_anteriores: DadosMensalAnterior,
                     caminho_saida: str,
                     config: ConfigFormulas,
                     filtrar_seco: bool = False,
                     areas_seco: Optional[List[int]] = None) -> pd.DataFrame:
        """Gera o novo arquivo mensal e retorna o DataFrame final"""
        self.log("=" * 60)
        self.log("🚀 INICIANDO GERAÇÃO DO MENSAL")
        self.log("=" * 60)
        self.progress(50)
        
        df = dados_anteriores.df.copy()
        
        if filtrar_seco and areas_seco:
            self.log(f"🔍 Aplicando filtro SECO (áreas: {areas_seco})")
            df = self._aplicar_filtro_seco(df, config, areas_seco)
            self.progress(60)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Mensal"
        
        self._aplicar_cabecalho(ws)
        self._escrever_dados_base(ws, df)
        self.progress(70)
        
        self._aplicar_formulas(ws, config, len(df))
        self._ajustar_larguras(ws)
        self.progress(80)
        
        os.makedirs(os.path.dirname(caminho_saida) or ".", exist_ok=True)
        wb.save(caminho_saida)
        self.progress(90)
        
        self.log("=" * 60)
        self.log(f"✅ ARQUIVO GERADO: {os.path.basename(caminho_saida)}")
        self.log(f"📊 Total de linhas: {len(df)}")
        self.log("=" * 60)
        self.progress(100)
        
        # Retorna DataFrame para visualização
        return df
    
    def _aplicar_filtro_seco(self, df: pd.DataFrame, config: ConfigFormulas, 
                            areas_seco: List[int]) -> pd.DataFrame:
        """Filtra apenas barras da região SECO"""
        mapeamento = self.carregar_mapeamento_area(config.parecer_file, config.mapping_sheet)
        
        if mapeamento is None:
            self.log("⚠ Filtro SECO não aplicado (mapeamento indisponível)")
            return df
        
        mapeamento['Área'] = pd.to_numeric(mapeamento['Área'], errors='coerce')
        barras_seco = mapeamento[mapeamento['Área'].isin(areas_seco)][['Nº Barra']]
        df_filtrado = df.merge(barras_seco, on='Nº Barra', how='inner')
        
        self.log(f"✓ Filtro SECO: {len(df)} -> {len(df_filtrado)} linhas")
        return df_filtrado
    
    def _aplicar_cabecalho(self, ws):
        """Aplica formatação ao cabeçalho"""
        headers = ["Nº Barra", "Nome", "Agente", "Área", "Questionamento SISBAR", "Parecer da Área"]
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=11)
        alignment = Alignment(horizontal="center", vertical="center")
        
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = header
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
    
    def _escrever_dados_base(self, ws, df: pd.DataFrame):
        """Escreve os dados base"""
        for _, row in df.iterrows():
            ws.append([
                row.get('Nº Barra', ""),
                row.get('Nome', ""),
                row.get('Agente', ""),
                "", "", ""
            ])
    
    def _aplicar_formulas(self, ws, config: ConfigFormulas, n_linhas: int):
        """Aplica as fórmulas do Excel"""
        self.log("📝 Aplicando fórmulas...")
        
        ref_parecer = Utils.excel_quote_external_ref(config.parecer_file, config.parecer_sheet_nome)
        ref_mapping = Utils.excel_quote_external_ref(config.parecer_file, config.mapping_sheet)
        
        sep = config.sep
        fallback_escaped = config.fallback_text.replace('"', '""')
        
        for row_idx in range(2, n_linhas + 2):
            cell_barra = f"$A${row_idx}"
            
            formula_area = f"=PROCV({cell_barra}{sep}{ref_mapping}${config.mapping_range}{sep}4{sep}FALSO)"
            ws.cell(row=row_idx, column=4).value = formula_area
            
            formula_quest = f"=SEERRO(PROCV({cell_barra}{sep}{ref_parecer}${config.parecer_range}{sep}5{sep}FALSO){sep}\"{fallback_escaped}\")"
            ws.cell(row=row_idx, column=5).value = formula_quest
            
            formula_parecer = f"=SEERRO(PROCV({cell_barra}{sep}{ref_parecer}${config.parecer_range}{sep}6{sep}FALSO){sep}\" \")"
            ws.cell(row=row_idx, column=6).value = formula_parecer
        
        self.log(f"✓ Fórmulas aplicadas em {n_linhas} linhas")
    
    def _ajustar_larguras(self, ws):
        """Ajusta largura das colunas"""
        larguras = [12, 40, 28, 8, 50, 50]
        for idx, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = largura


# ================================== GUI ===================================== #

class ChecklistWidget(QWidget):
    """Widget de checklist das etapas"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.checkboxes = {}
        etapas = [
            ("1", "Mensal Anterior Carregado"),
            ("2", "Arquivo de Parecer Selecionado"),
            ("3", "Valores A1 Lidos"),
            ("4", "Configurações Verificadas"),
            ("5", "Arquivo Gerado com Sucesso")
        ]
        
        for id_etapa, texto in etapas:
            cb = QCheckBox(f"✓ {texto}")
            cb.setEnabled(False)
            self.checkboxes[id_etapa] = cb
            layout.addWidget(cb)
        
        layout.addStretch()
    
    def marcar_etapa(self, id_etapa: str, completo: bool = True):
        """Marca uma etapa como completa"""
        if id_etapa in self.checkboxes:
            self.checkboxes[id_etapa].setChecked(completo)


class PreviewAbasWidget(QWidget):
    """Widget para preview dos valores A1 das abas"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Aba", "Valor A1"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        
        layout.addWidget(QLabel("📋 Preview dos Valores A1:"))
        layout.addWidget(self.table)
    
    def atualizar_preview(self, valores_a1: Dict[str, str]):
        """Atualiza a tabela de preview"""
        self.table.setRowCount(len(valores_a1))
        
        for row, (aba, valor) in enumerate(valores_a1.items()):
            item_aba = QTableWidgetItem(aba)
            item_valor = QTableWidgetItem(valor[:100] + "..." if len(valor) > 100 else valor)
            
            self.table.setItem(row, 0, item_aba)
            self.table.setItem(row, 1, item_valor)
        
        self.table.resizeColumnsToContents()


class TabelaFinalWidget(QWidget):
    """Widget para visualização da tabela final"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        self.lbl_info = QLabel("Aguardando geração...")
        
        layout.addWidget(self.lbl_info)
        layout.addWidget(self.table)
    
    def atualizar_tabela(self, df: pd.DataFrame):
        """Atualiza a tabela com o DataFrame"""
        if df is None or df.empty:
            return
        
        self.lbl_info.setText(f"📊 Tabela Final - {len(df)} linhas")
        
        # Configura tabela
        self.table.setRowCount(min(len(df), 100))  # Mostra primeiras 100 linhas
        self.table.setColumnCount(len(df.columns))
        self.table.setHorizontalHeaderLabels(list(df.columns))
        
        # Preenche dados
        for row in range(min(len(df), 100)):
            for col in range(len(df.columns)):
                valor = str(df.iloc[row, col])
                item = QTableWidgetItem(valor)
                self.table.setItem(row, col, item)
        
        self.table.resizeColumnsToContents()
        
        if len(df) > 100:
            self.lbl_info.setText(f"📊 Tabela Final - {len(df)} linhas (mostrando primeiras 100)")


class MainWindow(QWidget):
    """Janela principal com abas"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Automação SISBAR SECO - v3.0 [Com Preview e Checklist]")
        self.resize(1200, 800)
        
        self.processador = ProcessadorSISBAR(self._log, self._atualizar_progresso)
        self.dados_anteriores: Optional[DadosMensalAnterior] = None
        self.valores_a1: Dict[str, str] = {}
        self.df_final: Optional[pd.DataFrame] = None
        
        self._criar_interface()
        self._aplicar_estilos()
    
    def _criar_interface(self):
        """Cria a interface com abas"""
        layout_principal = QVBoxLayout(self)
        
        # Abas
        self.tabs = QTabWidget()
        
        # Aba 1: Configuração
        tab_config = self._criar_aba_configuracao()
        self.tabs.addTab(tab_config, "⚙️ Configuração")
        
        # Aba 2: Checklist
        self.checklist_widget = ChecklistWidget()
        self.tabs.addTab(self.checklist_widget, "✓ Checklist")
        
        # Aba 3: Preview Abas
        self.preview_widget = PreviewAbasWidget()
        self.tabs.addTab(self.preview_widget, "📋 Preview Abas")
        
        # Aba 4: Tabela Final
        self.tabela_widget = TabelaFinalWidget()
        self.tabs.addTab(self.tabela_widget, "📊 Tabela Final")
        
        # Aba 5: Logs
        tab_logs = QWidget()
        layout_logs = QVBoxLayout(tab_logs)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("Consolas", 9))
        layout_logs.addWidget(self.txt_log)
        self.tabs.addTab(tab_logs, "📝 Logs")
        
        layout_principal.addWidget(self.tabs)
        
        # Barra de progresso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout_principal.addWidget(self.progress_bar)
    
    def _criar_aba_configuracao(self) -> QWidget:
        """Cria a aba de configuração"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Grupo: Arquivos
        grupo_arq = QGroupBox("📁 Arquivos")
        layout_arq = QVBoxLayout()
        
        # Mensal Anterior
        h1 = QHBoxLayout()
        h1.addWidget(QLabel("Mensal Anterior:"))
        self.ed_mensal = QLineEdit()
        self.ed_mensal.setPlaceholderText("Ex: C:/Users/.../Mensal JAN26.xlsx")
        h1.addWidget(self.ed_mensal)
        btn_mensal = QPushButton("📂")
        btn_mensal.clicked.connect(self._escolher_mensal)
        h1.addWidget(btn_mensal)
        layout_arq.addLayout(h1)
        
        # Arquivo de Parecer
        h2 = QHBoxLayout()
        h2.addWidget(QLabel("Arquivo Parecer:"))
        self.ed_parecer = QLineEdit()
        self.ed_parecer.setPlaceholderText("Ex: C:/Users/.../Casos fev26.xlsx")
        h2.addWidget(self.ed_parecer)
        btn_parecer = QPushButton("📄")
        btn_parecer.clicked.connect(self._escolher_parecer)
        h2.addWidget(btn_parecer)
        layout_arq.addLayout(h2)
        
        # Arquivo de Saída
        h3 = QHBoxLayout()
        h3.addWidget(QLabel("Arquivo de Saída:"))
        self.ed_saida = QLineEdit()
        h3.addWidget(self.ed_saida)
        btn_sugerir = QPushButton("💡 Sugerir")
        btn_sugerir.clicked.connect(self._sugerir_saida)
        h3.addWidget(btn_sugerir)
        layout_arq.addLayout(h3)
        
        grupo_arq.setLayout(layout_arq)
        layout.addWidget(grupo_arq)
        
        # Grupo: Parecer
        grupo_parecer = QGroupBox("⚙️ Configuração de Parecer")
        layout_parecer = QVBoxLayout()
        
        h4 = QHBoxLayout()
        h4.addWidget(QLabel("Aba Parecer:"))
        self.ed_parecer_aba = QLineEdit("Parecer da Área Janeiro 2026")
        h4.addWidget(self.ed_parecer_aba)
        h4.addWidget(QLabel("Intervalo:"))
        self.ed_parecer_range = QLineEdit("A3:F70")
        self.ed_parecer_range.setMaximumWidth(100)
        h4.addWidget(self.ed_parecer_range)
        layout_parecer.addLayout(h4)
        
        h5 = QHBoxLayout()
        h5.addWidget(QLabel("Aba Mapeamento:"))
        self.ed_mapping_aba = QLineEdit("Parecer")
        h5.addWidget(self.ed_mapping_aba)
        h5.addWidget(QLabel("Intervalo:"))
        self.ed_mapping_range = QLineEdit("A2:D70")
        self.ed_mapping_range.setMaximumWidth(100)
        h5.addWidget(self.ed_mapping_range)
        layout_parecer.addLayout(h5)
        
        layout_parecer.addWidget(QLabel("Abas para ler A1:"))
        self.ed_abas_a1 = QLineEdit(", ".join(DEFAULT_ABAS))
        layout_parecer.addWidget(self.ed_abas_a1)
        
        h6 = QHBoxLayout()
        btn_ler_a1 = QPushButton("📖 Ler A1 das Abas")
        btn_ler_a1.clicked.connect(self._ler_a1)
        h6.addWidget(btn_ler_a1)
        h6.addWidget(QLabel("Aba fallback:"))
        self.cb_aba_fallback = QComboBox()
        h6.addWidget(self.cb_aba_fallback)
        layout_parecer.addLayout(h6)
        
        h7 = QHBoxLayout()
        h7.addWidget(QLabel("Separador:"))
        self.cb_sep = QComboBox()
        self.cb_sep.addItems([";", ","])
        self.cb_sep.setMaximumWidth(60)
        h7.addWidget(self.cb_sep)
        h7.addStretch()
        layout_parecer.addLayout(h7)
        
        grupo_parecer.setLayout(layout_parecer)
        layout.addWidget(grupo_parecer)
        
        # Grupo: SECO
        grupo_seco = QGroupBox("🔍 Filtro SECO")
        layout_seco = QVBoxLayout()
        
        self.chk_filtrar_seco = QCheckBox("Filtrar apenas barras da região SECO")
        self.chk_filtrar_seco.setChecked(True)
        layout_seco.addWidget(self.chk_filtrar_seco)
        
        h8 = QHBoxLayout()
        h8.addWidget(QLabel("Áreas SECO:"))
        self.ed_areas_seco = QLineEdit(",".join(map(str, AREAS_SECO_DEFAULT)))
        h8.addWidget(self.ed_areas_seco)
        layout_seco.addLayout(h8)
        
        grupo_seco.setLayout(layout_seco)
        layout.addWidget(grupo_seco)
        
        # Botão Gerar
        self.btn_gerar = QPushButton("🚀 GERAR MENSAL")
        self.btn_gerar.setMinimumHeight(50)
        self.btn_gerar.clicked.connect(self._on_gerar)
        layout.addWidget(self.btn_gerar)
        
        layout.addStretch()
        return widget
    
    def _aplicar_estilos(self):
        """Aplica estilos CSS"""
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4472C4;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                background-color: #4472C4;
                color: white;
                border: none;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #365899;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QTableWidget {
                border: 1px solid #ccc;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
                border: 2px solid #4CAF50;
            }
        """)
    
    # ---------------------- SLOTS ---------------------- #
    
    def _escolher_mensal(self):
        """Escolhe arquivo mensal anterior"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Escolha o Mensal Anterior", "", "Excel (*.xlsx)"
        )
        if arquivo:
            self.ed_mensal.setText(arquivo)
            self._carregar_mensal(arquivo)
    
    def _escolher_parecer(self):
        """Escolhe arquivo de parecer"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Escolha o Arquivo de Parecer", "", "Excel (*.xlsx)"
        )
        if arquivo:
            self.ed_parecer.setText(arquivo)
            self.checklist_widget.marcar_etapa("2")
    
    def _sugerir_saida(self):
        """Sugere nome do arquivo de saída"""
        if not self.dados_anteriores:
            QMessageBox.warning(self, "Atenção", "Carregue o mensal anterior primeiro")
            return
        
        nome_sugerido = Utils.gerar_nome_proximo_mensal(self.dados_anteriores.arquivo_origem)
        if nome_sugerido:
            pasta = os.path.dirname(self.dados_anteriores.arquivo_origem)
            self.ed_saida.setText(os.path.join(pasta, nome_sugerido))
            self._log(f"✓ Nome sugerido: {nome_sugerido}")
    
    def _carregar_mensal(self, caminho: str):
        """Carrega o arquivo mensal"""
        try:
            self.dados_anteriores = self.processador.carregar_mensal_anterior(caminho)
            self.checklist_widget.marcar_etapa("1")
            self._sugerir_saida()
            
            # Mostra preview na tabela final
            self.tabela_widget.atualizar_tabela(self.dados_anteriores.df)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao carregar mensal:\n{e}")
    
    def _ler_a1(self):
        """Lê valores A1 das abas"""
        parecer = self.ed_parecer.text().strip()
        if not parecer:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo de Parecer primeiro")
            return
        
        abas_texto = self.ed_abas_a1.text()
        abas = [a.strip() for a in abas_texto.split(',') if a.strip()]
        
        try:
            self.valores_a1 = self.processador.ler_valores_a1(parecer, abas)
            self.cb_aba_fallback.clear()
            self.cb_aba_fallback.addItems(abas)
            
            # Atualiza preview
            self.preview_widget.atualizar_preview(self.valores_a1)
            self.checklist_widget.marcar_etapa("3")
            
            # Muda para aba de preview
            self.tabs.setCurrentIndex(2)
            
            self._log(f"✓ Leitura A1 concluída ({len(abas)} abas)")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao ler A1:\n{e}")
    
    def _on_gerar(self):
        """Gera o arquivo mensal"""
        # Validações
        if not self.dados_anteriores:
            QMessageBox.warning(self, "Atenção", "Carregue o mensal anterior primeiro")
            return
        
        saida = self.ed_saida.text().strip()
        if not saida:
            QMessageBox.warning(self, "Atenção", "Defina o arquivo de saída")
            return
        
        parecer = self.ed_parecer.text().strip()
        if not parecer:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo de Parecer")
            return
        
        if not self.valores_a1:
            QMessageBox.warning(self, "Atenção", "Leia os valores A1 primeiro")
            return
        
        self.checklist_widget.marcar_etapa("4")
        
        # Configuração
        aba_fallback = self.cb_aba_fallback.currentText()
        texto_fallback = self.valores_a1.get(aba_fallback, "")
        
        config = ConfigFormulas(
            sep=self.cb_sep.currentText(),
            parecer_file=parecer,
            parecer_sheet_nome=self.ed_parecer_aba.text().strip(),
            parecer_range=self.ed_parecer_range.text().strip(),
            mapping_sheet=self.ed_mapping_aba.text().strip(),
            mapping_range=self.ed_mapping_range.text().strip(),
            fallback_aba=aba_fallback,
            fallback_text=texto_fallback
        )
        
        # Filtro SECO
        filtrar = self.chk_filtrar_seco.isChecked()
        areas_seco = None
        if filtrar:
            try:
                areas_texto = self.ed_areas_seco.text().strip()
                areas_seco = [int(a.strip()) for a in areas_texto.split(',') if a.strip()]
            except:
                QMessageBox.warning(self, "Atenção", "Áreas SECO inválidas")
                return
        
        # Processa
        self.btn_gerar.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        try:
            self.df_final = self.processador.gerar_mensal(
                self.dados_anteriores,
                saida,
                config,
                filtrar,
                areas_seco
            )
            
            # Atualiza tabela final
            self.tabela_widget.atualizar_tabela(self.df_final)
            self.checklist_widget.marcar_etapa("5")
            
            # Muda para aba da tabela final
            self.tabs.setCurrentIndex(3)
            
            QMessageBox.information(self, "Sucesso!", 
                f"Arquivo gerado com sucesso!\n\n{saida}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", 
                f"Falha ao gerar arquivo:\n\n{str(e)}\n\n{traceback.format_exc()}")
        finally:
            self.btn_gerar.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def _log(self, mensagem: str):
        """Adiciona mensagem ao log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.txt_log.append(f"[{timestamp}] {mensagem}")
    
    def _atualizar_progresso(self, valor: int):
        """Atualiza barra de progresso"""
        self.progress_bar.setValue(valor)
        QApplication.processEvents()


# ================================== MAIN ==================================== #

def main():
    """Função principal"""
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 9))
    
    janela = MainWindow()
    janela.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()