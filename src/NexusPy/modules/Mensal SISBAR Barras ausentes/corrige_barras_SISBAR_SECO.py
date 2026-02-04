

# 03/02/2026
# Versão: 1.1.3

# # Automação para corrigir as barras ausentes no SISBAR SECO 

# Arquivo de referência: "C:\Users\pedrovictor.veras\Downloads\Corrige_Barras_SECO_Fev26.xlsx"

# - Abrir o ultimo excel da pasta Mensal SISBAR Barras ausentes. Ex: Mensal - Jan26.xlsx
# - Pegar e acumular na planilha atual o Nº Barra, Nome, Agente, Área 
# - Colocar a nova planilha do mes consequentemente com o nome do mes e o ano autualizado. Ex: Mensal - Fev26.xlsx
# - Pegar pelo ID da barra e nome da barra com PROCV e procurar a resposta do "Parecer da Área" do caso do mês anterior

# - Pegar o parecer da área e verificar nas abas do excel: ["Futuras Ligadas", "Futuras Desligadas", "Desativadas","Ativas Desligadas 01 ", "Ativas Desligadas 02","Ativas Faltantes 01", "Ativas Faltantes 02"]
# - Pegar a resposta de cada Aba na cell A1 de cada aba do excel selecionada 
# - Formula para coluna "Questionamento SISBAR": =SEERRO(PROCV([@[Nº Barra]];'Parecer da Área Janeiro 2026'!$A$3:$F$70;5;FALSO); "<texto da cell A1 da aba selecionada>")
# - Formula para coluna "Parecer da Área": =SEERRO(PROCV([@[Nº Barra]];'Parecer da Área Janeiro 2026'!$A$3:$F$70;6;FALSO); " "))

# - Pegar o número da Área da região de cada barra com ID e filtrar as do SECO (Sudeste e Centro Oeste)
# - Formula 1 para coluna "Área": =PROCV([@[Nº Barra]];Parecer!$A$2:$D$70;4;FALSO)
# - Formula 2 para coluna "Área": =ESQUERDA(DIREITA(A1;5);2) -> Onde A1 é a cell do nome da barra




# app.py
# -*- coding: utf-8 -*-
"""
App GUI (PySide6) para automatizar a geração do arquivo "Mensal - <Mês><aa>.xlsx"

Requisitos no ambiente local do usuário:
    pip install pandas openpyxl PySide6

Observações importantes:
- O Excel é quem calcula as fórmulas. O openpyxl apenas grava as fórmulas.
- As fórmulas estão em PT-BR (SEERRO, PROCV, FALSO) e usam separador ";" por padrão.
- Referências externas em fórmulas exigem caminho + [arquivo.xlsx]aba!intervalo.
- Para fallback do "Questionamento SISBAR", o app permite escolher uma das abas
  e pegar o texto da célula A1 desta aba; o texto é gravado dentro da fórmula SEERRO.
- Opcionalmente é possível filtrar as barras para SECO antes de gravar (é necessário
  informar o arquivo e a aba que contém o mapeamento Nº Barra -> Área, além de
  quais números de Área pertencem ao SECO).

Classes principais:
- CorrigeBarrasSeco (controller): executa as operações de I/O, fórmulas e geração.
- PlannerMensalWidget (tela): interface PySide6 para o fluxo descrito.
"""

from __future__ import annotations
import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime # Adicionado
from pathlib import Path # Adicionado
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# GUI
try:
    from PySide6.QtWidgets import (
        QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
        QPushButton, QFileDialog, QTextEdit, QComboBox, QMessageBox, QCheckBox
    )
    from PySide6.QtCore import Qt
except Exception:
    # Permite que o arquivo seja importado sem ter PySide6 instalado
    QApplication = object  # type: ignore
    QWidget = object  # type: ignore
    QVBoxLayout = object  # type: ignore
    QHBoxLayout = object  # type: ignore
    QLabel = object  # type: ignore
    QLineEdit = object  # type: ignore
    QPushButton = object  # type: ignore
    QFileDialog = object  # type: ignore
    QTextEdit = object  # type: ignore
    QComboBox = object  # type: ignore
    QMessageBox = object  # type: ignore
    QCheckBox = object  # type: ignore
    Qt = object  # type: ignore


# ----------------------------- Utils & Config ----------------------------- #

PT_MONTHS = [
    "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
    "Jul", "Ago", "Set", "Out", "Nov", "Dez"
]


def logsafe(text: str) -> str:
    return text.replace("\n", " ")


def excel_quote_external_ref(path: str, sheet: str) -> str:
    """Monta o prefixo de referência externa para fórmulas Excel.
    Ex.: C:\\pasta\\[arquivo.xlsx]Aba -> 'C:\\pasta\\[arquivo.xlsx]Aba'!
    O retorno é algo como: 'C:\\pasta\\[arquivo.xlsx]Aba'!
    """
    # Excel aceita caminho com barra invertida. Se tiver aspas simples, duplica.
    norm = os.path.normpath(path)
    base = os.path.basename(norm)
    folder = os.path.dirname(norm)
    ref = f"{folder}\\[{base}]{sheet}"
    ref = ref.replace("'", "''")
    return f"'{ref}'!"


def next_month_code_from_filename(filename: str) -> Optional[str]:
    """Tenta extrair "Mensal - <Mon><yy>.xlsx" e sugerir o próximo mês.
    Retorna algo como "Mensal - Fev26.xlsx" ou None se não casar o padrão.
    """
    pat = r"(?i)mensal\s*-\s*([A-Za-zçãéêíóúÁÉÍÓÚÂÊÔÃÕÇ]{3})(\d{2})\.xlsx$"
    m = re.search(pat, os.path.basename(filename))
    if not m:
        return None
    mon_str = m.group(1).capitalize()
    yy = int(m.group(2))

    # Normaliza mês para índice
    try:
        idx = PT_MONTHS.index(mon_str[:3])
    except ValueError:
        return None

    next_idx = (idx + 1) % 12
    next_mon = PT_MONTHS[next_idx]
    next_yy = yy + 1 if next_idx == 0 else yy
    return f"Mensal - {next_mon}{next_yy:02d}.xlsx"


@dataclass
class FormulaContext:
    sep: str  # separador de argumentos em fórmulas ("," ou ";")
    parecer_file: Optional[str]  # arquivo que contém as abas de parecer (pode ser None -> usar interno)
    parecer_arearesp_sheet: str  # ex.: 'Parecer da Área Janeiro 2026'
    parecer_arearesp_range: str  # ex.: 'A3:F70'
    parecer_mapping_sheet: str   # ex.: 'Parecer'
    parecer_mapping_range: str   # ex.: 'A2:D70'
    fallback_text: str           # texto para SEERRO() na coluna Questionamento SISBAR


# -------------------------- Controller (Core) ----------------------------- #

class CorrigeBarrasSeco:
    """Controller com as operações de leitura, composição e gravação."""

    def __init__(self, logger_callable=print):
        self.log = logger_callable

    # ------------------------ Carregamento de dados ----------------------- #

    def detect_last_mensal(self, folder: str) -> Optional[str]:
        """Busca o arquivo mais recente que case com o padrão 'Mensal - <Mon><yy>.xlsx'
        com base na convenção de nomenclatura de mês/ano."""
        if not Path(folder).is_dir():
            self.log(f"Pasta inválida: {folder}")
            return None
        
        latest_file = None
        latest_date = None
        
        for file_path in Path(folder).glob("Mensal - *[0-9][0-9].xlsx"):
            match = re.search(r"Mensal - ([A-Za-z]+)(\d{2})\\.xlsx", file_path.name)
            if match:
                month_str, year_str = match.groups()
                # Mapear nome do mês para número para facilitar a comparação
                # Usar PT_MONTHS já definido
                try:
                    month_num = PT_MONTHS.index(month_str[:3].capitalize()) + 1
                except ValueError:
                    continue # Ignorar se o mês não for reconhecido

                try:
                    # Assumindo que o ano é 20YY, ou seja, 20 + year_str
                    file_date = datetime(2000 + int(year_str), month_num, 1)
                    if latest_date is None or file_date > latest_date:
                        latest_date = file_date
                        latest_file = file_path
                except ValueError:
                    # Ignorar arquivos com datas inválidas
                    pass
        
        if latest_file:
            self.log(f"Último mensal detectado: {latest_file.name}")
            return str(latest_file)
        else:
            self.log("Nenhum 'Mensal - *.xlsx' encontrado na pasta com convenção de nomenclatura válida.")
            return None

    def load_prev_mensal(self, file_path: str) -> pd.DataFrame:
        """Carrega o mensal anterior.
        Tenta detectar colunas: 'Nº Barra' (ou similares), 'Nome', 'Agente', 'Área' (se existir).
        """
        self.log(f"Lendo mensal anterior: {file_path}")
        # Lê a primeira planilha
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
        # Normaliza nomes de colunas
        colmap = self._detect_columns(df.columns)
        df2 = pd.DataFrame()
        for target, candidates in colmap.items():
            if candidates:
                df2[target] = df[candidates[0]]
            else:
                # Cria coluna vazia se não encontrada (ex.: Área pode ser por fórmula)
                df2[target] = ""
        n_before = len(df2)
        # Remove linhas totalmente vazias em Nº Barra
        df2 = df2[~df2['Nº Barra'].isna() & (df2['Nº Barra'] != "")]
        self.log(f"Linhas carregadas: {n_before} -> {len(df2)} (após limpar Nº Barra vazio)")
        return df2.reset_index(drop=True)

    def _detect_columns(self, cols: List[str]) -> Dict[str, List[str]]:
        """Retorna mapeamento dos nomes-padrão para os nomes reais encontrados.
        target keys: 'Nº Barra', 'Nome', 'Agente', 'Área'
        """
        lowered = {c.lower(): c for c in cols}
        def find(*options: str) -> List[str]:
            for opt in options:
                if opt.lower() in lowered:
                    return [lowered[opt.lower()]]
            # tenta aproximações simples
            for c in cols:
                cl = c.lower()
                if any(opt in cl for opt in options):  # substring
                    return [c]
            return []

        mapping = {
            'Nº Barra': find('nº barra', 'no barra', 'nºbarra', 'n barra', 'id barra', 'num barra', 'número barra', 'nro barra', 'nº'),
            'Nome': find('nome', 'nome barra'),
            'Agente': find('agente', 'empresa', 'concessionária', 'concessionaria'),
            'Área': find('área', 'area', 'região', 'regiao'),
        }
        return mapping

    def read_a1_values(self, file_path: str, sheets: List[str]) -> Dict[str, str]:
        """Lê os valores de A1 das abas informadas."""
        out: Dict[str, str] = {}
        self.log(f"Lendo A1 de {len(sheets)} aba(s) em: {file_path}")
        wb = load_workbook(file_path, data_only=True)
        for sh in sheets:
            if sh in wb.sheetnames:
                val = wb[sh]['A1'].value
                out[sh] = str(val) if val is not None else ""
            else:
                out[sh] = ""  # não encontrado
        return out

    def load_mapping_area(self, parecer_file: str, mapping_sheet: str) -> Optional[pd.DataFrame]:
        """Lê mapeamento Nº Barra -> Área a partir do arquivo de parecer."""
        try:
            df = pd.read_excel(parecer_file, sheet_name=mapping_sheet, engine='openpyxl')
            # tentativa de normalização simples
            colmap = {}
            for c in df.columns:
                cl = str(c).strip().lower()
                if 'barra' in cl and 'n' in cl:
                    colmap['Nº Barra'] = c
                elif 'área' in cl or 'area' in cl:
                    colmap['Área'] = c
            if 'Nº Barra' not in colmap or 'Área' not in colmap:
                self.log("Aviso: não foi possível detectar colunas de mapeamento (Nº Barra, Área).")
                return None
            return df[[colmap['Nº Barra'], colmap['Área']]].rename(columns={colmap['Nº Barra']: 'Nº Barra', colmap['Área']: 'Área'})
        except Exception as e:
            self.log(f"Falha ao ler mapeamento de área: {e}")
            return None

    # ------------------------ Escrita / Geração --------------------------- #

    def load_mapping_area(self, parecer_file: str, mapping_sheet: str) -> Optional[pd.DataFrame]:
        """Lê mapeamento Nº Barra -> Área a partir do arquivo de parecer."""
        try:
            df = pd.read_excel(parecer_file, sheet_name=mapping_sheet, engine='openpyxl')
            # tentativa de normalização simples
            colmap = {}
            for c in df.columns:
                cl = str(c).strip().lower()
                if 'barra' in cl and 'n' in cl:
                    colmap['Nº Barra'] = c
                elif 'área' in cl or 'area' in cl:
                    colmap['Área'] = c
            if 'Nº Barra' not in colmap or 'Área' not in colmap:
                self.log("Aviso: não foi possível detectar colunas de mapeamento (Nº Barra, Área).")
                return None
            return df[[colmap['Nº Barra'], colmap['Área']]].rename(columns={colmap['Nº Barra']: 'Nº Barra', colmap['Área']: 'Área'})
        except Exception as e:
            self.log(f"Falha ao ler mapeamento de área: {e}")
            return None

    def _load_parecer_area_data(self, file_path: str, sheet_name: str, data_range: str) -> Optional[pd.DataFrame]:
        """Lê os dados de parecer da área de um arquivo Excel e um intervalo específico."""
        try:
            # openpyxl ranges são 1-based, pandas é 0-based.
            # Para ler um range específico, pode-se usar read_excel com skiprows e nrows
            # Mas o mais fácil é carregar tudo e filtrar no pandas.
            # Ou, usar load_workbook para ler o range específico com openpyxl.
            wb = load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                self.log(f"Aba de parecer não encontrada: {sheet_name}")
                return None
            
            ws = wb[sheet_name]
            data = ws[data_range] # Retorna tupla de tuplas
            
            # Converter para lista de listas, ignorando células None
            rows_data = []
            for row_tuple in data:
                row_values = [cell.value for cell in row_tuple]
                rows_data.append(row_values)
            
            # A primeira linha do range é o cabeçalho se o range incluir cabeçalho
            # Considerando que A3:F70, então A3 é o primeiro dado, não o cabeçalho
            # Se o range incluir o cabeçalho, precisaríamos de uma lógica diferente.
            # Por simplicidade, assumirei que o range é apenas os dados.
            # Se A3:F70 contiver cabeçalho, precisaremos de um ajuste aqui.

            # Para PROCV, precisamos de um df com Nº Barra como a primeira coluna e as outras colunas
            # para as posições 5 e 6 (índices 4 e 5).
            # Os cabeçalhos não estão no range A3:F70. PROCV espera que o valor buscado
            # esteja na primeira coluna da matriz.
            # Vamos assumir que o range dado (A3:F70) já são os dados.

            # Nomes das colunas esperadas para o PROCV:
            # 1: Nº Barra (para PROCV)
            # ... (outras 3 colunas, que não importam para nós agora)
            # 5: Questionamento SISBAR (índice 4)
            # 6: Parecer da Área (índice 5)
            # Assumindo que a coluna Nº Barra é a primeira no range A3:F70
            # e que as colunas 5 e 6 são as que precisamos.

            # Se o range for A3:F70 e a coluna A é Nº Barra, B, C, D, E é 5, F é 6.
            # Precisamos do Nº Barra, e das colunas E e F.

            # Transformar rows_data em DataFrame. Nomes das colunas serão genéricos a princípio.
            # Depois, podemos renomear ou acessar por índice.
            df_parecer = pd.DataFrame(rows_data, columns=[f'Col{i+1}' for i in range(len(rows_data[0]))])

            # As colunas que nos interessam são a 1 (Nº Barra), 5 e 6.
            # Se o range for A3:F70, então A é Col1, E é Col5, F é Col6.
            # Renomear para clareza:
            df_parecer = df_parecer.rename(columns={'Col1': 'Nº Barra', 'Col5': 'Questionamento SISBAR Col', 'Col6': 'Parecer da Área Col'})
            
            # Definir Nº Barra como índice para buscas rápidas (similar ao PROCV com FALSO)
            df_parecer.set_index('Nº Barra', inplace=True)
            return df_parecer

        except Exception as e:
            self.log(f"Falha ao ler dados de parecer da área: {e}")
            return None

    def generate_output(self,
                        df_base: pd.DataFrame,
                        output_path: str,
                        ctx: FormulaContext,
                        filter_seco: bool = False,
                        seco_areas: Optional[List[int]] = None) -> None:
        """Gera o arquivo de saída, escreve dados e fórmulas.
        - df_base deve conter colunas: ['Nº Barra','Nome','Agente','Área'] (Área pode estar vazio)
        - ctx define origem das fórmulas e texto fallback.
        - se filter_seco=True, tenta filtrar por áreas em seco_areas usando arquivo de parecer (mapeamento).
        """
        self.log(f"Gerando: {output_path}")
        df = df_base.copy()

        # Carrega dados de parecer para cálculos em Python
        parecer_df = None
        if ctx.parecer_file:
            parecer_df = self._load_parecer_area_data(ctx.parecer_file, ctx.parecer_arearesp_sheet, ctx.parecer_arearesp_range)
            if parecer_df is None:
                self.log("Aviso: Dados de parecer da área não puderam ser carregados para cálculos em Python.")
        
        # Adicionar colunas para os resultados dos cálculos
        df['Questionamento SISBAR'] = ''
        df['Parecer da Área'] = ''
        
        # Aplicar lógica para "Questionamento SISBAR"
        # O [Nº Barra] do df_base é a chave para lookup no parecer_df
        if parecer_df is not None:
            # Usar .get para simular PROCV com FALSO e SEERRO
            df['Questionamento SISBAR'] = df['Nº Barra'].apply(
                lambda num_barra: parecer_df.get(num_barra, {}).get('Questionamento SISBAR Col', ctx.fallback_text) if pd.notna(num_barra) else ctx.fallback_text
            )
        else:
            # Se não houver parecer_df, preencher com o fallback_text
            df['Questionamento SISBAR'] = ctx.fallback_text

        # Aplicar lógica para "Parecer da Área"
        if parecer_df is not None:
            df['Parecer da Área'] = df['Nº Barra'].apply(
                lambda num_barra: parecer_df.get(num_barra, {}).get('Parecer da Área Col', " ") if pd.notna(num_barra) else " "
            )
        else:
            df['Parecer da Área'] = " "

        # Carregar mapeamento de área
        mapping_df = None
        if ctx.parecer_file:
            mapping_df = self.load_mapping_area(ctx.parecer_file, ctx.parecer_mapping_sheet)
            if mapping_df is not None:
                mapping_df.set_index('Nº Barra', inplace=True)
            else:
                self.log("Aviso: Mapeamento de área não puderam ser carregados para cálculos em Python.")

        # Aplicar lógica para coluna "Área" (PROCV com Fallback)
        # Se a coluna Área já tiver valores, preservá-los, senão aplicar fórmula.
        df['Área'] = df.apply(lambda row: 
            row['Área'] if pd.notna(row['Área']) and row['Área'] != '' else (
                str(mapping_df.get(row['Nº Barra'], {}).get('Área', '')) if mapping_df is not None and pd.notna(row['Nº Barra']) else (
                    re.search(r"\b(\w{2})\d{3}$", str(row['Nome']))[1] if re.search(r"\b(\w{2})\d{3}$", str(row['Nome'])) else ''
                )
            ),
            axis=1
        )

        # Pré-filtro SECO (opcional) usando o arquivo de parecer/mapeamento
        if filter_seco and seco_areas and ctx.parecer_file:
            mapping = self.load_mapping_area(ctx.parecer_file, ctx.parecer_mapping_sheet)
            if mapping is not None:
                self.log(f"Aplicando filtro SECO. Áreas: {seco_areas}")
                mapping['Área'] = pd.to_numeric(mapping['Área'], errors='coerce')
                keep = mapping[mapping['Área'].isin(seco_areas)][['Nº Barra']]
                df = df.merge(keep, on='Nº Barra', how='inner')
                self.log(f"Linhas após filtro SECO: {len(df)}")
            else:
                self.log("Filtro SECO não aplicado: mapeamento indisponível.")

        # Cria workbook e aba
        wb = Workbook()
        ws = wb.active
        ws.title = "Mensal"

        # Cabeçalhos
        headers = ["Nº Barra", "Nome", "Agente", "Área", "Questionamento SISBAR", "Parecer da Área"]
        ws.append(headers)

        # Escreve valores base nas 3 primeiras colunas (e Área se já vier)
        for _, row in df.iterrows():
            ws.append([
                row.get('Nº Barra', ""),
                row.get('Nome', ""),
                row.get('Agente', ""),
                row.get('Área', ""),  # pode estar vazio; será formula depois
                "",  # placeholder para fórmula
                ""   # placeholder para fórmula
            ])

        # Ajuste de largura simples
        widths = [12, 40, 28, 8, 50, 50]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        # Escreve os dados do DataFrame no worksheet
        for r_idx, row_data in df.iterrows():
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx + 2, column=c_idx, value=value) # +2 porque cabeçalho é linha 1, pandas é 0-based


        # Salva arquivo
        # Garante diretório
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        self.log(f"Arquivo gerado: {output_path}")


# ------------------------------ UI (PySide6) ------------------------------ #

class PlannerMensalWidget(QWidget):
    """Tela principal para configurar e executar o fluxo."""

    def __init__(self):  # type: ignore
        super().__init__()
        self.setWindowTitle("Automação - Corrige Barras SECO (Mensal)")
        self.controller = CorrigeBarrasSeco(self._log)
        self._build_ui()
        # Tentar detectar o último mensal ao iniciar
        assets_path = os.path.join(os.path.dirname(__file__), "assets")
        last_mensal = self.controller.detect_last_mensal(assets_path)
        if last_mensal:
            self.ed_prev.setText(last_mensal)
            # Sugerir nome de saída com base no anterior, se detectado
            sug = next_month_code_from_filename(last_mensal)
            if sug:
                out_dir = os.path.dirname(last_mensal)
                self.ed_out.setText(os.path.join(out_dir, sug))
        # Tentar detectar o último mensal ao iniciar
        assets_path = os.path.join(os.path.dirname(__file__), "assets")
        last_mensal = self.controller.detect_last_mensal(assets_path)
        if last_mensal:
            self.ed_prev.setText(last_mensal)

    # ---------------------------- UI Helpers ------------------------------ #

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Arquivo Mensal anterior
        self.ed_prev = QLineEdit()
        btn_prev = QPushButton("Escolher Mensal anterior…")
        btn_prev.clicked.connect(self._choose_prev)
        btn_detect_prev = QPushButton("Detectar Último Mensal") # Novo botão
        btn_detect_prev.clicked.connect(self._detect_prev_mensal) # Nova conexão

        # Sugerir nome de saída com base no anterior
        self.ed_out = QLineEdit()
        btn_suggest = QPushButton("Sugerir próximo nome")
        btn_suggest.clicked.connect(self._suggest_output_name)

        # Arquivo com PARECER (onde estão as abas e mapeamentos)
        self.ed_parecer = QLineEdit()
        btn_parecer = QPushButton("Escolher arquivo de Parecer…")
        btn_parecer.clicked.connect(self._choose_parecer)

        # Aba com respostas (A3:F70) e intervalo
        self.ed_arearesp_sheet = QLineEdit("Parecer da Área Janeiro 2026")
        self.ed_arearesp_range = QLineEdit("A3:F70")

        # Aba com mapeamento Nº Barra -> Área (A2:D70) e intervalo
        self.ed_mapping_sheet = QLineEdit("Parecer")
        self.ed_mapping_range = QLineEdit("A2:D70")

        # Lista de abas para capturar A1 e seleção de uma aba para fallback
        self.ed_abas = QLineEdit("Futuras Ligadas, Futuras Desligadas, Desativadas, Ativas Desligadas 01, Ativas Desligadas 02, Ativas Faltantes 01, Ativas Faltantes 02")
        btn_ler_a1 = QPushButton("Ler A1 das abas…")
        btn_ler_a1.clicked.connect(self._ler_a1)
        self.cb_aba_fallback = QComboBox()
        self.lbl_a1_preview = QLabel("A1 selecionado: (vazio)")
        self.cb_aba_fallback.currentTextChanged.connect(self._update_a1_preview)

        # Separador de fórmula
        self.cb_sep = QComboBox()
        self.cb_sep.addItems([";", ","])  # padrão ;

        # Filtro SECO (opcional)
        self.chk_seco = QCheckBox("Filtrar apenas SECO")
        self.ed_seco_areas = QLineEdit("")
        self.ed_seco_areas.setPlaceholderText("Informe números de Área separados por vírgula. Ex.: 10,11,12")

        # Botão GERAR
        btn_gerar = QPushButton("Gerar Mensal")
        btn_gerar.clicked.connect(self._on_generate)

        # Log
        self.log = QTextEdit()
        self.log.setReadOnly(True)

        # Layouts
        def row(label: str, widget):
            box = QHBoxLayout()
            box.addWidget(QLabel(label))
            if isinstance(widget, (QLineEdit, QComboBox, QLabel)):
                box.addWidget(widget)
            else:
                for w in widget:
                    box.addWidget(w)
            lay.addLayout(box)

        row("Mensal anterior:", [self.ed_prev, btn_prev, btn_detect_prev]) # Adicionado btn_detect_prev
        row("Arquivo de saída:", [self.ed_out, btn_suggest])
        row("Arquivo de Parecer:", [self.ed_parecer, btn_parecer])
        row("Aba 'Parecer da Área':", self.ed_arearesp_sheet)
        row("Intervalo 'Parecer da Área':", self.ed_arearesp_range)
        row("Aba mapeamento (Nº Barra->Área):", self.ed_mapping_sheet)
        row("Intervalo mapeamento:", self.ed_mapping_range)
        row("Abas (A1) separadas por vírgula:", self.ed_abas)
        row("Ações A1:", [btn_ler_a1, self.cb_aba_fallback])
        row("Prévia A1:", self.lbl_a1_preview)
        row("Separador de fórmula:", self.cb_sep)
        row("Filtro SECO?", [self.chk_seco])
        row("Áreas SECO:", self.ed_seco_areas)

        lay.addWidget(btn_gerar)
        lay.addWidget(QLabel("Log:"))
        lay.addWidget(self.log)

    # ---------------------------- Log helper ------------------------------ #

    def _log(self, msg: str):
        if isinstance(self.log, QTextEdit):
            self.log.append(msg)
        else:
            print(msg)

    # ------------------------ UI Event Handlers --------------------------- #

    def _detect_prev_mensal(self):
        assets_path = os.path.join(os.path.dirname(__file__), "assets")
        last_mensal = self.controller.detect_last_mensal(assets_path)
        if last_mensal:
            self.ed_prev.setText(last_mensal)
        else:
            QMessageBox.information(self, "Info", "Nenhum 'Mensal - *.xlsx' encontrado na pasta assets.")

    def _choose_prev(self):
        path, _ = QFileDialog.getOpenFileName(self, "Escolha o Mensal anterior", "", "Excel (*.xlsx)")
        if path:
            self.ed_prev.setText(path)

    def _choose_parecer(self):
        path, _ = QFileDialog.getOpenFileName(self, "Escolha o arquivo de Parecer", "", "Excel (*.xlsx)")
        if path:
            self.ed_parecer.setText(path)

    def _suggest_output_name(self):
        prev = self.ed_prev.text().strip()
        if not prev:
            QMessageBox.warning(self, "Atenção", "Selecione o Mensal anterior primeiro.")
            return
        sug = next_month_code_from_filename(prev)
        if not sug:
            QMessageBox.information(self, "Info", "Não foi possível inferir o próximo nome pelo padrão do arquivo anterior.")
        else:
            out_dir = os.path.dirname(prev)
            self.ed_out.setText(os.path.join(out_dir, sug))

    def _ler_a1(self):
        parecer = self.ed_parecer.text().strip()
        if not parecer:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo de Parecer para ler as abas.")
            return
        abas = [a.strip() for a in self.ed_abas.text().split(',') if a.strip()]
        try:
            a1 = self.controller.read_a1_values(parecer, abas)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao ler A1: {e}")
            return
        self.cb_aba_fallback.clear()
        self.cb_aba_fallback.addItems(abas)
        # salva como atributo para prévia
        self._a1_cache = a1
        self._update_a1_preview()
        self._log(f"Abas lidas: {len(abas)}. Você pode escolher a aba para fallback.")

    def _update_a1_preview(self):
        aba = self.cb_aba_fallback.currentText() if hasattr(self, 'cb_aba_fallback') else ''
        val = getattr(self, '_a1_cache', {}).get(aba, '') if hasattr(self, '_a1_cache') else ''
        self.lbl_a1_preview.setText(f"A1 selecionado: {val}")

    def _on_generate(self):
        try:
            prev = self.ed_prev.text().strip()
            out_path = self.ed_out.text().strip()
            parecer = self.ed_parecer.text().strip()
            arearesp_sheet = self.ed_arearesp_sheet.text().strip()
            arearesp_range = self.ed_arearesp_range.text().strip()
            mapping_sheet = self.ed_mapping_sheet.text().strip()
            mapping_range = self.ed_mapping_range.text().strip()
            sep = self.cb_sep.currentText()

            if not prev or not out_path:
                QMessageBox.warning(self, "Atenção", "Informe o Mensal anterior e o arquivo de saída.")
                return

            # Carrega base do mensal anterior
            df = self.controller.load_prev_mensal(prev)
            if df.empty:
                QMessageBox.warning(self, "Atenção", "Mensal anterior não possui linhas válidas de 'Nº Barra'.")
                return

            # Fallback text a partir da aba escolhida
            fallback = ""
            if hasattr(self, '_a1_cache') and isinstance(getattr(self, '_a1_cache'), dict):
                aba_sel = self.cb_aba_fallback.currentText()
                fallback = getattr(self, '_a1_cache', {}).get(aba_sel, "")

            ctx = FormulaContext(
                sep=sep,
                parecer_file=parecer if parecer else None,
                parecer_arearesp_sheet=arearesp_sheet,
                parecer_arearesp_range=arearesp_range,
                parecer_mapping_sheet=mapping_sheet,
                parecer_mapping_range=mapping_range,
                fallback_text=fallback or ""
            )

            # Filtro SECO (opcional)
            do_filter = self.chk_seco.isChecked()
            seco_areas = None
            if do_filter:
                txt = self.ed_seco_areas.text().strip()
                seco_areas = []
                if txt:
                    for p in txt.split(','):
                        p = p.strip()
                        if p:
                            try:
                                seco_areas.append(int(p))
                            except Exception:
                                pass
                if not seco_areas:
                    QMessageBox.warning(self, "Atenção", "Informe ao menos um número de Área para o filtro SECO.")
                    return

            self.controller.generate_output(
                df_base=df,
                output_path=out_path,
                ctx=ctx,
                filter_seco=do_filter,
                seco_areas=seco_areas
            )
            QMessageBox.information(self, "Sucesso", f"Arquivo gerado:\n{out_path}")
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Erro", f"Falha ao gerar: {e}")


# ------------------------------ Entry Point ------------------------------ #

def main():
    app = QApplication(sys.argv)

    # Estilo personalizado (QSS)
    app.setStyleSheet("""
        QWidget {
            background-color: #f0f0f0;
            font-family: Arial, sans-serif;
            font-size: 10pt;
        }
        QLabel {
            color: #333333;
        }
        QLineEdit, QTextEdit, QComboBox {
            border: 1px solid #cccccc;
            border-radius: 4px;
            padding: 5px;
            background-color: #ffffff;
            color: #333333;
        }
        QPushButton {
            background-color: #007bff;
            color: white;
            border: none;
            border-radius: 4px;
            padding: 8px 15px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: #0056b3;
        }
        QMessageBox {
            background-color: #ffffff;
            color: #333333;
        }
        QMessageBox QLabel {
            color: #333333;
        }
    """)

    w = PlannerMensalWidget()
    w.resize(980, 720)
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
