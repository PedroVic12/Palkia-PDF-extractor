# -*- coding: utf-8 -*-
"""
Automação SISBAR SECO - Correção de Barras Ausentes
Versão: 4.0.0
Data: 11/02/2026

Sistema completo para automatizar a geração do arquivo "Mensal - <Mês><aa>.xlsx"
com todas as fórmulas Excel necessárias.

Requisitos:
    pip install pandas openpyxl PySide6

Funcionalidades:
- Detecta automaticamente o último arquivo mensal na pasta
- Acumula dados do mês anterior (No., Nome, Agente, Área)
- Gera novo arquivo com nome do mês/ano atualizado
- Aplica fórmulas PROCV para buscar parecer do mês anterior
- Lê valores A1 das abas de classificação
- Aplica fórmulas Excel corretas para todas as colunas
- Filtro opcional para região SECO
"""

from __future__ import annotations
import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# GUI
try:
    from PySide6.QtWidgets import (
        QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
        QPushButton, QFileDialog, QTextEdit, QComboBox, QMessageBox, 
        QCheckBox, QGroupBox, QTabWidget, QTableWidget, 
        QTableWidgetItem, QHeaderView, QProgressBar
    )
    from PySide6.QtCore import Qt
    from PySide6.QtGui import QFont
except ImportError:
    print("ERRO: PySide6 não instalado. Execute: pip install PySide6")
    sys.exit(1)


# ================================ CONSTANTES ================================ #

PT_MONTHS = [
    "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
    "Jul", "Ago", "Set", "Out", "Nov", "Dez"
]

DEFAULT_ABAS = [
    "Futuras Ligadas",
    "Futuras Desligadas", 
    "Desativadas",
    "Ativas Desligadas 01",
    "Ativas Desligadas 02",
    "Ativas Faltantes 01",
    "Ativas Faltantes 02"
]

AREAS_SECO_DEFAULT = [10, 11, 12, 20, 21, 22, 30, 31, 32]


# ================================ DATACLASSES =============================== #

@dataclass
class ConfigFormulas:
    """Configurações para geração de fórmulas Excel"""
    sep: str  # Separador de argumentos (; ou ,)
    parecer_file: str  # Caminho do arquivo de parecer
    parecer_sheet_nome: str  # Nome da aba com parecer (ex: "Parecer da Área Janeiro 2026")
    parecer_range: str  # Intervalo (ex: "A3:F70")
    mapping_sheet: str  # Aba com mapeamento No. -> Área (ex: "Parecer")
    mapping_range: str  # Intervalo do mapeamento (ex: "A2:D70")
    fallback_text: str  # Texto fallback para Questionamento SISBAR (vem do A1 da aba selecionada)


@dataclass
class DadosMensalAnterior:
    """Dados extraídos do mensal anterior"""
    df: pd.DataFrame
    arquivo_origem: str
    mes_ano_origem: Tuple[str, str]


# ================================ UTILITÁRIOS =============================== #

class Utils:
    """Funções utilitárias"""
    
    @staticmethod
    def extrair_mes_ano(filename: str) -> Optional[Tuple[str, str]]:
        """Extrai mês e ano de um nome de arquivo"""
        # Padrão mais flexível para capturar nomes de meses
        pattern = r"(?i)mensal\s*[_-]?\s*([A-Za-zçãéêíóúÁÉÍÓÚÂÊÔÃÕÇ]+)(\d{2})\.xlsx$"
        match = re.search(pattern, os.path.basename(filename))
        if match:
            mes_str = match.group(1).strip()
            ano_str = match.group(2)
            
            # Normaliza o nome do mês para os 3 primeiros caracteres
            mes_str = mes_str[:3].capitalize()
            
            # Mapeia variações comuns
            mes_map = {
                "Mar": "Mar", "Março": "Mar", "Marco": "Mar",
                "Jan": "Jan", "Janeiro": "Jan",
                "Fev": "Fev", "Fevereiro": "Fev",
                "Abr": "Abr", "Abril": "Abr",
                "Mai": "Mai", "Maio": "Mai",
                "Jun": "Jun", "Junho": "Jun",
                "Jul": "Jul", "Julho": "Jul",
                "Ago": "Ago", "Agosto": "Ago",
                "Set": "Set", "Setembro": "Set",
                "Out": "Out", "Outubro": "Out",
                "Nov": "Nov", "Novembro": "Nov",
                "Dez": "Dez", "Dezembro": "Dez"
            }
            
            # Tenta encontrar no mapeamento
            for key, value in mes_map.items():
                if mes_str.lower().startswith(key.lower()[:3]):
                    return (value, ano_str)
            
            # Se não encontrou, retorna os 3 primeiros caracteres
            return (mes_str[:3].capitalize(), ano_str)
        return None
    
    @staticmethod
    def calcular_proximo_mes(mes: str, ano: str) -> Tuple[str, str]:
        """Calcula o próximo mês/ano"""
        try:
            idx = PT_MONTHS.index(mes)
            proximo_idx = (idx + 1) % 12
            proximo_mes = PT_MONTHS[proximo_idx]
            proximo_ano = str(int(ano) + 1) if proximo_idx == 0 else ano
            return (proximo_mes, proximo_ano)
        except ValueError:
            raise ValueError(f"Mês inválido: {mes}")
    
    @staticmethod
    def gerar_nome_proximo_mensal(arquivo_anterior: str) -> Optional[str]:
        """Gera o nome do próximo arquivo mensal"""
        mes_ano = Utils.extrair_mes_ano(arquivo_anterior)
        if not mes_ano:
            return None
        
        mes, ano = mes_ano
        proximo_mes, proximo_ano = Utils.calcular_proximo_mes(mes, ano)
        return f"Mensal - {proximo_mes}{proximo_ano}.xlsx"
    
    @staticmethod
    def excel_quote_external_ref(caminho: str, aba: str) -> str:
        """Cria referência externa para fórmula Excel
        
        Retorna algo como: 'C:\\pasta\\[arquivo.xlsx]Aba'!
        """
        norm = os.path.normpath(caminho)
        base = os.path.basename(norm)
        folder = os.path.dirname(norm)
        ref = f"{folder}\\[{base}]{aba}"
        ref = ref.replace("'", "''")
        return f"'{ref}'!"
    
    @staticmethod
    def detectar_ultimo_mensal(pasta: str) -> Optional[str]:
        """Detecta o último arquivo mensal na pasta baseado na data do nome do arquivo"""
        if not Path(pasta).is_dir():
            return None
        
        arquivos_encontrados = []
        
        for arquivo_path in Path(pasta).glob("Mensal*.xlsx"):
            nome = arquivo_path.name
            mes_ano = Utils.extrair_mes_ano(nome)
            if mes_ano:
                mes, ano = mes_ano
                try:
                    idx_mes = PT_MONTHS.index(mes)
                    # Converte para data para comparação
                    data_arquivo = datetime(2000 + int(ano), idx_mes + 1, 1)
                    arquivos_encontrados.append((data_arquivo, str(arquivo_path)))
                except (ValueError, IndexError):
                    continue
        
        if not arquivos_encontrados:
            return None
        
        # Retorna o arquivo com a data mais recente
        arquivos_encontrados.sort(key=lambda x: x[0], reverse=True)
        return arquivos_encontrados[0][1]


# ============================== PROCESSADOR ================================= #

class ProcessadorSISBAR:
    """Classe principal para processar os dados SISBAR"""
    
    def __init__(self, logger_fn=print, progress_fn=None):
        self.log = logger_fn
        self.progress = progress_fn or (lambda x: None)
    
    def carregar_mensal_anterior(self, caminho: str) -> DadosMensalAnterior:
        """Carrega dados do mensal anterior"""
        self.log(f"📖 Lendo: {os.path.basename(caminho)}")
        self.progress(10)
        
        try:
            # Primeiro, tenta ler com openpyxl diretamente para ver o que tem
            wb = load_workbook(caminho, data_only=False)  # data_only=False para pegar fórmulas também
            ws = wb.active
            self.log(f"  Aba ativa: {ws.title}")
            self.log(f"  Dimensões da planilha: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Lê valores brutos das primeiras linhas para debug
            if ws.max_row > 0:
                primeira_linha_valores = []
                for col in range(1, min(ws.max_column + 1, 7)):
                    cell = ws.cell(row=1, column=col)
                    primeira_linha_valores.append(f"{cell.coordinate}={cell.value}")
                self.log(f"  Primeira linha (cabeçalho): {primeira_linha_valores}")
                
                if ws.max_row > 1:
                    segunda_linha_valores = []
                    for col in range(1, min(ws.max_column + 1, 7)):
                        cell = ws.cell(row=2, column=col)
                        segunda_linha_valores.append(f"{cell.coordinate}={cell.value}")
                    self.log(f"  Segunda linha (dados?): {segunda_linha_valores}")
            
            wb.close()
            
            # Agora tenta ler com pandas - primeiro com data_only=True (valores calculados)
            try:
                df_raw = pd.read_excel(caminho, sheet_name=0, engine='openpyxl')
                self.log(f"  Colunas encontradas (pandas): {list(df_raw.columns)}")
                self.log(f"  Total de linhas brutas (pandas): {len(df_raw)}")
            except Exception as e:
                self.log(f"  ⚠ Erro ao ler com pandas (data_only=True): {e}")
                df_raw = pd.DataFrame()
            
            # Se não encontrou dados, tenta ler diretamente do openpyxl
            if len(df_raw) == 0:
                self.log("  Tentando ler diretamente do openpyxl...")
                wb = load_workbook(caminho, data_only=False)  # Primeiro tenta sem calcular fórmulas
                ws = wb.active
                
                # Encontra cabeçalho (pode estar na linha 1 ou 2)
                headers = []
                header_row = 1
                
                # Verifica se linha 1 tem cabeçalhos válidos
                primeira_linha = []
                for col in range(1, min(ws.max_column + 1, 10)):
                    cell = ws.cell(row=1, column=col)
                    primeira_linha.append(str(cell.value) if cell.value else "")
                
                # Se primeira linha parece ter cabeçalhos, usa ela
                if any("barra" in str(v).lower() or "nome" in str(v).lower() for v in primeira_linha):
                    header_row = 1
                else:
                    # Tenta linha 2
                    segunda_linha = []
                    for col in range(1, min(ws.max_column + 1, 10)):
                        cell = ws.cell(row=2, column=col)
                        segunda_linha.append(str(cell.value) if cell.value else "")
                    if any("barra" in str(v).lower() or "nome" in str(v).lower() for v in segunda_linha):
                        header_row = 2
                
                # Lê cabeçalhos
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=header_row, column=col)
                    if cell.value:
                        headers.append(str(cell.value).strip())
                    else:
                        headers.append(f"Col{col}")
                
                self.log(f"  Cabeçalhos encontrados na linha {header_row}: {headers[:6]}")
                
                # Lê dados linha por linha (começa após o cabeçalho)
                dados = []
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    linha = []
                    tem_dados = False
                    valor_barra = None
                    
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        valor = cell.value
                        
                        # Se é fórmula, pega a fórmula como string
                        if cell.data_type == 'f':
                            valor = cell.value  # Mantém a fórmula
                            if valor:
                                tem_dados = True
                        elif valor is not None:
                            valor_str = str(valor).strip()
                            if valor_str != "" and valor_str.lower() != "nan":
                                tem_dados = True
                                # Se é a primeira coluna (No.), guarda o valor
                                if col_idx == 1:
                                    valor_barra = valor_str
                        
                        linha.append(valor)
                    
                    # Só adiciona se tiver dados E se a primeira coluna não estiver vazia
                    if tem_dados and valor_barra and valor_barra.lower() not in ["", "nan", "none"]:
                        dados.append(linha)
                
                wb.close()
                
                if dados:
                    self.log(f"  ✓ Linhas de dados encontradas (openpyxl): {len(dados)}")
                    # Garante que todas as linhas tenham o mesmo número de colunas
                    max_cols = max(len(linha) for linha in dados) if dados else len(headers)
                    dados_padronizados = []
                    for linha in dados:
                        linha_padronizada = linha + [None] * (max_cols - len(linha))
                        dados_padronizados.append(linha_padronizada[:len(headers)])
                    
                    df_raw = pd.DataFrame(dados_padronizados, columns=headers[:len(dados_padronizados[0])])
                else:
                    self.log("  ⚠ Nenhuma linha de dados encontrada após o cabeçalho")
                    # Tenta ler com data_only=True (valores calculados)
                    self.log("  Tentando ler com valores calculados (data_only=True)...")
                    try:
                        wb2 = load_workbook(caminho, data_only=True)
                        ws2 = wb2.active
                        dados_calc = []
                        for row_idx in range(header_row + 1, ws2.max_row + 1):
                            linha = []
                            tem_dados = False
                            for col_idx in range(1, len(headers) + 1):
                                cell = ws2.cell(row=row_idx, column=col_idx)
                                valor = cell.value
                                if valor is not None and str(valor).strip() != "":
                                    tem_dados = True
                                linha.append(valor)
                            if tem_dados:
                                dados_calc.append(linha)
                        wb2.close()
                        
                        if dados_calc:
                            self.log(f"  ✓ Linhas encontradas com valores calculados: {len(dados_calc)}")
                            max_cols = max(len(linha) for linha in dados_calc) if dados_calc else len(headers)
                            dados_padronizados = []
                            for linha in dados_calc:
                                linha_padronizada = linha + [None] * (max_cols - len(linha))
                                dados_padronizados.append(linha_padronizada[:len(headers)])
                            df_raw = pd.DataFrame(dados_padronizados, columns=headers[:len(dados_padronizados[0])])
                        else:
                            df_raw = pd.DataFrame(columns=headers)
                    except Exception as e2:
                        self.log(f"  ⚠ Erro ao ler com data_only=True: {e2}")
                        df_raw = pd.DataFrame(columns=headers)
            
            mapeamento = self._detectar_colunas(df_raw.columns)
            self.log(f"  Mapeamento detectado: {mapeamento}")
            
            df = pd.DataFrame()
            for col_padrao, col_encontrada in mapeamento.items():
                if col_encontrada:
                    df[col_padrao] = df_raw[col_encontrada[0]]
                    self.log(f"  ✓ Coluna '{col_padrao}' mapeada de '{col_encontrada[0]}'")
                else:
                    df[col_padrao] = ""
                    self.log(f"  ⚠ Coluna '{col_padrao}' não encontrada")
            
            n_antes = len(df)
            # Remove linhas onde No. está vazio ou é NaN
            # Converte para string para verificar melhor
            df['No.'] = df['No.'].astype(str)
            df = df[df['No.'].notna() & (df['No.'] != "") & (df['No.'].str.strip() != "") & (df['No.'] != "nan")]
            df = df.reset_index(drop=True)
            
            self.log(f"✓ Linhas carregadas: {len(df)} (de {n_antes})")
            if len(df) == 0:
                self.log("=" * 60)
                self.log("⚠⚠⚠ ATENÇÃO: DataFrame está vazio! ⚠⚠⚠")
                self.log("=" * 60)
                self.log("  Possíveis causas:")
                self.log("    1. Arquivo não tem dados (apenas cabeçalhos)")
                self.log("    2. Dados estão em formato de fórmulas não calculadas")
                self.log("    3. Coluna 'No.' está vazia em todas as linhas")
                self.log("    4. Arquivo foi gerado anteriormente e está vazio")
                self.log("")
                self.log("  SOLUÇÃO:")
                self.log("    - Abra o arquivo Excel manualmente e verifique se há dados")
                self.log("    - Se houver fórmulas, salve o arquivo novamente para calcular")
                self.log("    - Use um arquivo mensal anterior que tenha dados reais")
                self.log("=" * 60)
                
                # Mostra primeiras linhas para debug
                if n_antes > 0:
                    self.log(f"  Primeiras linhas do arquivo original:")
                    for idx, row in df_raw.head(5).iterrows():
                        valores_linha = {k: str(v)[:50] if v is not None else "None" for k, v in dict(row).items()}
                        self.log(f"    Linha {idx}: {valores_linha}")
                else:
                    self.log("  Nenhuma linha de dados encontrada no arquivo")
            
            self.progress(20)
            
            mes_ano = Utils.extrair_mes_ano(caminho)
            if not mes_ano:
                mes_ano = ("", "")
            
            return DadosMensalAnterior(
                df=df,
                arquivo_origem=caminho,
                mes_ano_origem=mes_ano
            )
        except Exception as e:
            self.log(f"❌ Erro ao carregar mensal: {e}")
            self.log(traceback.format_exc())
            raise
    
    def _detectar_colunas(self, colunas: List[str]) -> Dict[str, List[str]]:
        """Detecta colunas do Excel"""
        # Normaliza colunas removendo espaços extras e caracteres especiais
        colunas_normalizadas = {c.strip(): c for c in colunas}
        lower_map = {c.lower().strip(): c for c in colunas}
        
        def encontrar(*opcoes: str) -> List[str]:
            # Primeiro tenta match exato (case insensitive)
            for opcao in opcoes:
                opcao_lower = opcao.lower().strip()
                if opcao_lower in lower_map:
                    return [lower_map[opcao_lower]]
            
            # Depois tenta substring match
            for col in colunas:
                col_lower = col.lower().strip()
                for opt in opcoes:
                    opt_lower = opt.lower().strip()
                    # Match exato ou substring significativa
                    if opt_lower == col_lower or (len(opt_lower) >= 3 and opt_lower in col_lower):
                        return [col]
            
            return []
        
        resultado = {
            'No.': encontrar('no.', 'nº barra', 'no barra', 'num barra', 'id barra', 'nº', 'n barra', 'número barra', 'nro barra', 'barra'),
            'Nome': encontrar('nome', 'nome barra', 'nome da barra'),
            'Agente': encontrar('agente', 'empresa', 'concessionária', 'concessionaria'),
            'Área': encontrar('área', 'area', 'região', 'regiao'),
        }
        
        # Log de debug
        self.log(f"  Detecção de colunas:")
        for col_padrao, col_encontrada in resultado.items():
            if col_encontrada:
                self.log(f"    '{col_padrao}' -> '{col_encontrada[0]}'")
            else:
                self.log(f"    '{col_padrao}' -> NÃO ENCONTRADA")
        
        return resultado
    
    def ler_valores_a1(self, arquivo: str, abas: List[str]) -> Dict[str, str]:
        """Lê valores da célula A1 de cada aba"""
        self.log(f"📋 Lendo A1 de {len(abas)} aba(s)...")
        self.progress(30)
        resultado = {}
        
        try:
            wb = load_workbook(arquivo, data_only=True)
            for aba in abas:
                if aba in wb.sheetnames:
                    valor = wb[aba]['A1'].value
                    resultado[aba] = str(valor) if valor else ""
                    self.log(f"  • {aba}: '{resultado[aba][:50]}...'")
                else:
                    resultado[aba] = ""
                    self.log(f"  ⚠ {aba}: não encontrada")
            wb.close()
            self.progress(40)
        except Exception as e:
            self.log(f"❌ Erro ao ler A1: {e}")
            return {}
        
        return resultado
    
    def carregar_mapeamento_area(self, arquivo: str, aba: str) -> Optional[pd.DataFrame]:
        """Carrega mapeamento No. -> Área"""
        try:
            df = pd.read_excel(arquivo, sheet_name=aba, engine='openpyxl')
            
            col_barra = None
            col_area = None
            
            for col in df.columns:
                col_lower = str(col).lower().strip()
                # Prioriza "No." ou "no."
                if col_lower == 'no.' or col_lower.startswith('no.'):
                    col_barra = col
                elif 'barra' in col_lower and ('n' in col_lower or 'num' in col_lower or 'número' in col_lower):
                    if col_barra is None:  # Só usa se não encontrou "No."
                        col_barra = col
                elif 'área' in col_lower or 'area' in col_lower:
                    col_area = col
            
            if not col_barra or not col_area:
                self.log("⚠ Colunas de mapeamento não encontradas")
                return None
            
            df_map = df[[col_barra, col_area]].rename(
                columns={col_barra: 'No.', col_area: 'Área'}
            )
            
            self.log(f"✓ Mapeamento carregado: {len(df_map)} linhas")
            return df_map
            
        except Exception as e:
            self.log(f"❌ Erro ao carregar mapeamento: {e}")
            return None
    
    def gerar_preview(self,
                     dados_anteriores: DadosMensalAnterior,
                     config: ConfigFormulas,
                     filtrar_seco: bool = False,
                     areas_seco: Optional[List[int]] = None) -> pd.DataFrame:
        """Gera preview do DataFrame final sem salvar arquivo"""
        self.log("=" * 60)
        self.log("👁️ GERANDO PREVIEW")
        self.log("=" * 60)
        
        df = dados_anteriores.df.copy()
        
        if len(df) == 0:
            self.log("⚠ ATENÇÃO: DataFrame está vazio!")
            return df
        
        if filtrar_seco and areas_seco:
            self.log(f"🔍 Aplicando filtro SECO (áreas: {areas_seco})")
            df = self._aplicar_filtro_seco(df, config, areas_seco)
        
        # Adiciona colunas que serão preenchidas por fórmulas (para preview)
        df['Área'] = df.get('Área', '')
        df['Questionamento SISBAR'] = ''
        df['Parecer da Área'] = ''
        
        self.log(f"📊 Preview gerado: {len(df)} linhas")
        self.log("=" * 60)
        
        return df
    
    def gerar_mensal(self,
                     dados_anteriores: DadosMensalAnterior,
                     caminho_saida: str,
                     config: ConfigFormulas,
                     filtrar_seco: bool = False,
                     areas_seco: Optional[List[int]] = None) -> pd.DataFrame:
        """Gera o novo arquivo mensal e retorna o DataFrame final"""
        self.log("=" * 60)
        self.log("🚀 INICIANDO GERAÇÃO DO MENSAL")
        self.log("=" * 60)
        self.progress(50)
        
        df = dados_anteriores.df.copy()
        
        if len(df) == 0:
            raise ValueError("DataFrame está vazio! Não é possível gerar o arquivo.")
        
        if filtrar_seco and areas_seco:
            self.log(f"🔍 Aplicando filtro SECO (áreas: {areas_seco})")
            df = self._aplicar_filtro_seco(df, config, areas_seco)
            self.progress(60)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Mensal"
        
        self._aplicar_cabecalho(ws)
        self._escrever_dados_base(ws, df)
        self.progress(70)
        
        self._aplicar_formulas(ws, config, len(df))
        self._ajustar_larguras(ws)
        self.progress(80)
        
        os.makedirs(os.path.dirname(caminho_saida) or ".", exist_ok=True)
        wb.save(caminho_saida)
        self.progress(90)
        
        self.log("=" * 60)
        self.log(f"✅ ARQUIVO GERADO: {os.path.basename(caminho_saida)}")
        self.log(f"📊 Total de linhas: {len(df)}")
        self.log("=" * 60)
        self.progress(100)
        
        return df
    
    def _aplicar_filtro_seco(self, df: pd.DataFrame, config: ConfigFormulas, 
                            areas_seco: List[int]) -> pd.DataFrame:
        """Filtra apenas barras da região SECO"""
        mapeamento = self.carregar_mapeamento_area(config.parecer_file, config.mapping_sheet)
        
        if mapeamento is None:
            self.log("⚠ Filtro SECO não aplicado (mapeamento indisponível)")
            return df
        
        mapeamento['Área'] = pd.to_numeric(mapeamento['Área'], errors='coerce')
        barras_seco = mapeamento[mapeamento['Área'].isin(areas_seco)][['No.']]
        df_filtrado = df.merge(barras_seco, on='No.', how='inner')
        
        self.log(f"✓ Filtro SECO: {len(df)} -> {len(df_filtrado)} linhas")
        return df_filtrado
    
    def _aplicar_cabecalho(self, ws):
        """Aplica formatação ao cabeçalho"""
        headers = ["No.", "Nome", "Agente", "Área", "Questionamento SISBAR", "Parecer da Área"]
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=11)
        alignment = Alignment(horizontal="center", vertical="center")
        
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = header
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
    
    def _escrever_dados_base(self, ws, df: pd.DataFrame):
        """Escreve os dados base (No., Nome, Agente)"""
        for _, row in df.iterrows():
            ws.append([
                row.get('No.', ""),
                row.get('Nome', ""),
                row.get('Agente', ""),
                "",  # Área será preenchida por fórmula
                "",  # Questionamento SISBAR será preenchido por fórmula
                ""   # Parecer da Área será preenchido por fórmula
            ])
    
    def _aplicar_formulas(self, ws, config: ConfigFormulas, n_linhas: int):
        """Aplica as fórmulas Excel nas colunas apropriadas"""
        self.log("📝 Aplicando fórmulas...")
        
        # Cria referências externas para as fórmulas
        ref_parecer = Utils.excel_quote_external_ref(config.parecer_file, config.parecer_sheet_nome)
        ref_mapping = Utils.excel_quote_external_ref(config.parecer_file, config.mapping_sheet)
        
        sep = config.sep
        fallback_escaped = config.fallback_text.replace('"', '""')
        
        # Aplica fórmulas linha por linha
        for row_idx in range(2, n_linhas + 2):  # Começa na linha 2 (linha 1 é cabeçalho)
            cell_barra = f"$A${row_idx}"
            cell_nome = f"$B${row_idx}"
            
            # Fórmula 1: Área (PROCV com fallback usando ESQUERDA(DIREITA...))
            # Primeiro tenta PROCV, se falhar usa ESQUERDA(DIREITA(nome, 5), 2)
            formula_area_procv = f"PROCV({cell_barra}{sep}{ref_mapping}${config.mapping_range}{sep}4{sep}FALSO)"
            formula_area_fallback = f"ESQUERDA(DIREITA({cell_nome}{sep}5){sep}2)"
            formula_area = f"=SEERRO({formula_area_procv}{sep}{formula_area_fallback})"
            ws.cell(row=row_idx, column=4).value = formula_area
            
            # Fórmula 2: Questionamento SISBAR
            # =SEERRO(PROCV([@[No.]];'Parecer da Área Janeiro 2026'!$A$3:$F$70;5;FALSO); "<texto da cell A1 da aba selecionada>")
            formula_quest_procv = f"PROCV({cell_barra}{sep}{ref_parecer}${config.parecer_range}{sep}5{sep}FALSO)"
            formula_quest = f"=SEERRO({formula_quest_procv}{sep}\"{fallback_escaped}\")"
            ws.cell(row=row_idx, column=5).value = formula_quest
            
            # Fórmula 3: Parecer da Área
            # =SEERRO(PROCV([@[No.]];'Parecer da Área Janeiro 2026'!$A$3:$F$70;6;FALSO); " ")
            formula_parecer_procv = f"PROCV({cell_barra}{sep}{ref_parecer}${config.parecer_range}{sep}6{sep}FALSO)"
            formula_parecer = f"=SEERRO({formula_parecer_procv}{sep}\" \")"
            ws.cell(row=row_idx, column=6).value = formula_parecer
        
        self.log(f"✓ Fórmulas aplicadas em {n_linhas} linhas")
    
    def _ajustar_larguras(self, ws):
        """Ajusta largura das colunas"""
        larguras = [12, 40, 28, 8, 50, 50]
        for idx, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = largura


# ================================== GUI ===================================== #

class ChecklistWidget(QWidget):
    """Widget de checklist das etapas"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.checkboxes = {}
        etapas = [
            ("1", "Mensal Anterior Carregado"),
            ("2", "Arquivo de Parecer Selecionado"),
            ("3", "Valores A1 Lidos"),
            ("4", "Configurações Verificadas"),
            ("5", "Arquivo Gerado com Sucesso")
        ]
        
        for id_etapa, texto in etapas:
            cb = QCheckBox(f"✓ {texto}")
            cb.setEnabled(False)
            self.checkboxes[id_etapa] = cb
            layout.addWidget(cb)
        
        layout.addStretch()
    
    def marcar_etapa(self, id_etapa: str, completo: bool = True):
        """Marca uma etapa como completa"""
        if id_etapa in self.checkboxes:
            self.checkboxes[id_etapa].setChecked(completo)


class PreviewAbasWidget(QWidget):
    """Widget para preview dos valores A1 das abas"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Aba", "Valor A1"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        
        layout.addWidget(QLabel("📋 Preview dos Valores A1:"))
        layout.addWidget(self.table)
    
    def atualizar_preview(self, valores_a1: Dict[str, str]):
        """Atualiza a tabela de preview"""
        self.table.setRowCount(len(valores_a1))
        
        for row, (aba, valor) in enumerate(valores_a1.items()):
            item_aba = QTableWidgetItem(aba)
            item_valor = QTableWidgetItem(valor[:100] + "..." if len(valor) > 100 else valor)
            
            self.table.setItem(row, 0, item_aba)
            self.table.setItem(row, 1, item_valor)
        
        self.table.resizeColumnsToContents()


class TabelaFinalWidget(QWidget):
    """Widget para visualização da tabela final"""
    
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        self.lbl_info = QLabel("Aguardando geração...")
        
        layout.addWidget(self.lbl_info)
        layout.addWidget(self.table)
    
    def atualizar_tabela(self, df: pd.DataFrame):
        """Atualiza a tabela com o DataFrame"""
        if df is None or df.empty:
            self.lbl_info.setText("⚠️ Tabela vazia - Nenhum dado para exibir")
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return
        
        self.lbl_info.setText(f"📊 Tabela Final - {len(df)} linhas")
        
        # Configura tabela
        n_linhas_exibir = min(len(df), 100)  # Mostra primeiras 100 linhas
        self.table.setRowCount(n_linhas_exibir)
        self.table.setColumnCount(len(df.columns))
        self.table.setHorizontalHeaderLabels(list(df.columns))
        
        # Preenche dados
        for row in range(n_linhas_exibir):
            for col in range(len(df.columns)):
                valor = df.iloc[row, col]
                # Trata valores NaN/None
                if pd.isna(valor):
                    valor_str = ""
                else:
                    valor_str = str(valor)
                
                item = QTableWidgetItem(valor_str)
                self.table.setItem(row, col, item)
        
        self.table.resizeColumnsToContents()
        
        if len(df) > 100:
            self.lbl_info.setText(f"📊 Tabela Final - {len(df)} linhas (mostrando primeiras 100)")


class MainWindow(QWidget):
    """Janela principal com abas"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Automação SISBAR SECO - v4.0 [Versão Corrigida]")
        self.resize(1200, 800)
        
        self.processador = ProcessadorSISBAR(self._log, self._atualizar_progresso)
        self.dados_anteriores: Optional[DadosMensalAnterior] = None
        self.valores_a1: Dict[str, str] = {}
        self.df_final: Optional[pd.DataFrame] = None
        
        self._criar_interface()
        self._aplicar_estilos()
        self._detectar_mensal_automatico()
    
    def _criar_interface(self):
        """Cria a interface com abas"""
        layout_principal = QVBoxLayout(self)
        
        # Abas
        self.tabs = QTabWidget()
        
        # Aba 1: Configuração
        tab_config = self._criar_aba_configuracao()
        self.tabs.addTab(tab_config, "⚙️ Configuração")
        
        # Aba 2: Checklist
        self.checklist_widget = ChecklistWidget()
        self.tabs.addTab(self.checklist_widget, "✓ Checklist")
        
        # Aba 3: Preview Abas
        self.preview_widget = PreviewAbasWidget()
        self.tabs.addTab(self.preview_widget, "📋 Preview Abas")
        
        # Aba 4: Tabela Final
        self.tabela_widget = TabelaFinalWidget()
        self.tabs.addTab(self.tabela_widget, "📊 Tabela Final")
        
        # Aba 5: Logs
        tab_logs = QWidget()
        layout_logs = QVBoxLayout(tab_logs)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("Consolas", 9))
        layout_logs.addWidget(self.txt_log)
        self.tabs.addTab(tab_logs, "📝 Logs")
        
        layout_principal.addWidget(self.tabs)
        
        # Barra de progresso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout_principal.addWidget(self.progress_bar)
    
    def _criar_aba_configuracao(self) -> QWidget:
        """Cria a aba de configuração"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Grupo: Arquivos
        grupo_arq = QGroupBox("📁 Arquivos")
        layout_arq = QVBoxLayout()
        
        # Mensal Anterior
        h1 = QHBoxLayout()
        h1.addWidget(QLabel("Mensal Anterior:"))
        self.ed_mensal = QLineEdit()
        self.ed_mensal.setPlaceholderText("Ex: C:/Users/.../Mensal JAN26.xlsx")
        h1.addWidget(self.ed_mensal)
        btn_mensal = QPushButton("📂")
        btn_mensal.clicked.connect(self._escolher_mensal)
        h1.addWidget(btn_mensal)
        btn_detectar = QPushButton("🔍 Detectar")
        btn_detectar.clicked.connect(self._detectar_mensal_automatico)
        h1.addWidget(btn_detectar)
        layout_arq.addLayout(h1)
        
        # Arquivo de Parecer
        h2 = QHBoxLayout()
        h2.addWidget(QLabel("Arquivo Parecer:"))
        self.ed_parecer = QLineEdit()
        self.ed_parecer.setPlaceholderText("Ex: C:/Users/.../Casos fev26.xlsx")
        h2.addWidget(self.ed_parecer)
        btn_parecer = QPushButton("📄")
        btn_parecer.clicked.connect(self._escolher_parecer)
        h2.addWidget(btn_parecer)
        layout_arq.addLayout(h2)
        
        # Arquivo de Saída
        h3 = QHBoxLayout()
        h3.addWidget(QLabel("Arquivo de Saída:"))
        self.ed_saida = QLineEdit()
        h3.addWidget(self.ed_saida)
        btn_sugerir = QPushButton("💡 Sugerir")
        btn_sugerir.clicked.connect(self._sugerir_saida)
        h3.addWidget(btn_sugerir)
        layout_arq.addLayout(h3)
        
        grupo_arq.setLayout(layout_arq)
        layout.addWidget(grupo_arq)
        
        # Grupo: Parecer
        grupo_parecer = QGroupBox("⚙️ Configuração de Parecer")
        layout_parecer = QVBoxLayout()
        
        h4 = QHBoxLayout()
        h4.addWidget(QLabel("Aba Parecer:"))
        self.ed_parecer_aba = QLineEdit("Parecer da Área Janeiro 2026")
        h4.addWidget(self.ed_parecer_aba)
        h4.addWidget(QLabel("Intervalo:"))
        self.ed_parecer_range = QLineEdit("A3:F70")
        self.ed_parecer_range.setMaximumWidth(100)
        h4.addWidget(self.ed_parecer_range)
        layout_parecer.addLayout(h4)
        
        h5 = QHBoxLayout()
        h5.addWidget(QLabel("Aba Mapeamento:"))
        self.ed_mapping_aba = QLineEdit("Parecer")
        h5.addWidget(self.ed_mapping_aba)
        h5.addWidget(QLabel("Intervalo:"))
        self.ed_mapping_range = QLineEdit("A2:D70")
        self.ed_mapping_range.setMaximumWidth(100)
        h5.addWidget(self.ed_mapping_range)
        layout_parecer.addLayout(h5)
        
        layout_parecer.addWidget(QLabel("Abas para ler A1:"))
        self.ed_abas_a1 = QLineEdit(", ".join(DEFAULT_ABAS))
        layout_parecer.addWidget(self.ed_abas_a1)
        
        h6 = QHBoxLayout()
        btn_ler_a1 = QPushButton("📖 Ler A1 das Abas")
        btn_ler_a1.clicked.connect(self._ler_a1)
        h6.addWidget(btn_ler_a1)
        h6.addWidget(QLabel("Aba fallback:"))
        self.cb_aba_fallback = QComboBox()
        h6.addWidget(self.cb_aba_fallback)
        layout_parecer.addLayout(h6)
        
        h7 = QHBoxLayout()
        h7.addWidget(QLabel("Separador:"))
        self.cb_sep = QComboBox()
        self.cb_sep.addItems([";", ","])
        self.cb_sep.setMaximumWidth(60)
        h7.addWidget(self.cb_sep)
        h7.addStretch()
        layout_parecer.addLayout(h7)
        
        grupo_parecer.setLayout(layout_parecer)
        layout.addWidget(grupo_parecer)
        
        # Grupo: SECO
        grupo_seco = QGroupBox("🔍 Filtro SECO")
        layout_seco = QVBoxLayout()
        
        self.chk_filtrar_seco = QCheckBox("Filtrar apenas barras da região SECO")
        self.chk_filtrar_seco.setChecked(True)
        layout_seco.addWidget(self.chk_filtrar_seco)
        
        h8 = QHBoxLayout()
        h8.addWidget(QLabel("Áreas SECO:"))
        self.ed_areas_seco = QLineEdit(",".join(map(str, AREAS_SECO_DEFAULT)))
        h8.addWidget(self.ed_areas_seco)
        layout_seco.addLayout(h8)
        
        grupo_seco.setLayout(layout_seco)
        layout.addWidget(grupo_seco)
        
        # Botões de ação
        h_botoes = QHBoxLayout()
        self.btn_preview = QPushButton("👁️ PREVIEW")
        self.btn_preview.setMinimumHeight(50)
        self.btn_preview.clicked.connect(self._on_preview)
        h_botoes.addWidget(self.btn_preview)
        
        self.btn_gerar = QPushButton("🚀 GERAR MENSAL")
        self.btn_gerar.setMinimumHeight(50)
        self.btn_gerar.clicked.connect(self._on_gerar)
        h_botoes.addWidget(self.btn_gerar)
        
        layout.addLayout(h_botoes)
        
        layout.addStretch()
        return widget
    
    def _aplicar_estilos(self):
        """Aplica estilos CSS"""
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4472C4;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                background-color: #4472C4;
                color: white;
                border: none;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #365899;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QTableWidget {
                border: 1px solid #ccc;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
                border: 2px solid #4CAF50;
            }
        """)
    
    # ---------------------- SLOTS ---------------------- #
    
    def _detectar_mensal_automatico(self):
        """Detecta automaticamente o último mensal na pasta assets"""
        pasta_assets = os.path.join(os.path.dirname(__file__), "assets")
        ultimo_mensal = Utils.detectar_ultimo_mensal(pasta_assets)
        
        if ultimo_mensal:
            self.ed_mensal.setText(ultimo_mensal)
            self._carregar_mensal(ultimo_mensal)
            self._log(f"✓ Mensal detectado: {os.path.basename(ultimo_mensal)}")
        else:
            QMessageBox.information(self, "Info", 
                f"Nenhum arquivo mensal encontrado em:\n{pasta_assets}")
    
    def _escolher_mensal(self):
        """Escolhe arquivo mensal anterior"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Escolha o Mensal Anterior", "", "Excel (*.xlsx)"
        )
        if arquivo:
            self.ed_mensal.setText(arquivo)
            self._carregar_mensal(arquivo)
    
    def _escolher_parecer(self):
        """Escolhe arquivo de parecer"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Escolha o Arquivo de Parecer", "", "Excel (*.xlsx)"
        )
        if arquivo:
            self.ed_parecer.setText(arquivo)
            self.checklist_widget.marcar_etapa("2")
    
    def _sugerir_saida(self):
        """Sugere nome do arquivo de saída"""
        if not self.dados_anteriores:
            QMessageBox.warning(self, "Atenção", "Carregue o mensal anterior primeiro")
            return
        
        nome_sugerido = Utils.gerar_nome_proximo_mensal(self.dados_anteriores.arquivo_origem)
        if nome_sugerido:
            pasta = os.path.dirname(self.dados_anteriores.arquivo_origem)
            self.ed_saida.setText(os.path.join(pasta, nome_sugerido))
            self._log(f"✓ Nome sugerido: {nome_sugerido}")
    
    def _carregar_mensal(self, caminho: str):
        """Carrega o arquivo mensal"""
        try:
            self.dados_anteriores = self.processador.carregar_mensal_anterior(caminho)
            
            if len(self.dados_anteriores.df) == 0:
                QMessageBox.warning(
                    self, 
                    "Arquivo Vazio", 
                    "⚠️ O arquivo carregado está vazio!\n\n"
                    "O arquivo não contém dados válidos na coluna 'No.'.\n\n"
                    "Possíveis causas:\n"
                    "• Arquivo tem apenas cabeçalhos\n"
                    "• Dados estão em fórmulas não calculadas\n"
                    "• Arquivo foi gerado anteriormente e está vazio\n\n"
                    "SOLUÇÃO:\n"
                    "• Abra o arquivo Excel manualmente e verifique\n"
                    "• Use um arquivo mensal anterior com dados reais\n"
                    "• Se houver fórmulas, salve o arquivo para calcular"
                )
                self.checklist_widget.marcar_etapa("1", False)
            else:
                self.checklist_widget.marcar_etapa("1")
            
            self._sugerir_saida()
            
            # Mostra preview na tabela final
            self.tabela_widget.atualizar_tabela(self.dados_anteriores.df)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao carregar mensal:\n{e}")
            self.checklist_widget.marcar_etapa("1", False)
    
    def _ler_a1(self):
        """Lê valores A1 das abas"""
        parecer = self.ed_parecer.text().strip()
        if not parecer:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo de Parecer primeiro")
            return
        
        abas_texto = self.ed_abas_a1.text()
        abas = [a.strip() for a in abas_texto.split(',') if a.strip()]
        
        try:
            self.valores_a1 = self.processador.ler_valores_a1(parecer, abas)
            self.cb_aba_fallback.clear()
            self.cb_aba_fallback.addItems(abas)
            
            # Atualiza preview
            self.preview_widget.atualizar_preview(self.valores_a1)
            self.checklist_widget.marcar_etapa("3")
            
            # Muda para aba de preview
            self.tabs.setCurrentIndex(2)
            
            self._log(f"✓ Leitura A1 concluída ({len(abas)} abas)")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao ler A1:\n{e}")
    
    def _obter_config(self) -> Optional[ConfigFormulas]:
        """Obtém configuração validada"""
        parecer = self.ed_parecer.text().strip()
        if not parecer:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo de Parecer")
            return None
        
        if not self.valores_a1:
            QMessageBox.warning(self, "Atenção", "Leia os valores A1 primeiro")
            return None
        
        aba_fallback = self.cb_aba_fallback.currentText()
        texto_fallback = self.valores_a1.get(aba_fallback, "")
        
        return ConfigFormulas(
            sep=self.cb_sep.currentText(),
            parecer_file=parecer,
            parecer_sheet_nome=self.ed_parecer_aba.text().strip(),
            parecer_range=self.ed_parecer_range.text().strip(),
            mapping_sheet=self.ed_mapping_aba.text().strip(),
            mapping_range=self.ed_mapping_range.text().strip(),
            fallback_text=texto_fallback
        )
    
    def _obter_filtro_seco(self) -> Tuple[bool, Optional[List[int]]]:
        """Obtém configuração do filtro SECO"""
        filtrar = self.chk_filtrar_seco.isChecked()
        areas_seco = None
        if filtrar:
            try:
                areas_texto = self.ed_areas_seco.text().strip()
                areas_seco = [int(a.strip()) for a in areas_texto.split(',') if a.strip()]
            except:
                QMessageBox.warning(self, "Atenção", "Áreas SECO inválidas")
                return False, None
        return filtrar, areas_seco
    
    def _on_preview(self):
        """Gera preview sem salvar arquivo"""
        # Validações
        if not self.dados_anteriores:
            QMessageBox.warning(self, "Atenção", "Carregue o mensal anterior primeiro")
            return
        
        if len(self.dados_anteriores.df) == 0:
            QMessageBox.warning(
                self, 
                "Arquivo Vazio", 
                "⚠️ Não é possível gerar preview!\n\n"
                "O arquivo mensal anterior está vazio.\n\n"
                "Por favor:\n"
                "• Carregue um arquivo mensal que tenha dados\n"
                "• Ou verifique se o arquivo atual tem dados válidos"
            )
            return
        
        config = self._obter_config()
        if not config:
            return
        
        filtrar, areas_seco = self._obter_filtro_seco()
        if filtrar and areas_seco is None:
            return
        
        # Gera preview
        self.btn_preview.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        try:
            self.df_final = self.processador.gerar_preview(
                self.dados_anteriores,
                config,
                filtrar,
                areas_seco
            )
            
            # Atualiza tabela final
            self.tabela_widget.atualizar_tabela(self.df_final)
            
            # Muda para aba da tabela final
            self.tabs.setCurrentIndex(3)
            
            if len(self.df_final) == 0:
                QMessageBox.warning(
                    self, 
                    "Preview Vazio", 
                    "⚠️ Preview gerado mas está vazio!\n\n"
                    "Possíveis causas:\n"
                    "• O filtro SECO removeu todas as linhas\n"
                    "• O mensal anterior não tem dados válidos\n\n"
                    "Verifique:\n"
                    "• Se o arquivo mensal tem dados\n"
                    "• Se as áreas SECO estão corretas\n"
                    "• Se o mapeamento de área está funcionando"
                )
            else:
                QMessageBox.information(
                    self, 
                    "Preview Gerado", 
                    f"✅ Preview gerado com sucesso!\n\n"
                    f"Total de linhas: {len(self.df_final)}\n\n"
                    f"Revise a tabela na aba 'Tabela Final' antes de gerar o arquivo."
                )
        except Exception as e:
            QMessageBox.critical(self, "Erro", 
                f"Falha ao gerar preview:\n\n{str(e)}\n\n{traceback.format_exc()}")
        finally:
            self.btn_preview.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def _on_gerar(self):
        """Gera o arquivo mensal"""
        # Validações
        if not self.dados_anteriores:
            QMessageBox.warning(self, "Atenção", "Carregue o mensal anterior primeiro")
            return
        
        if len(self.dados_anteriores.df) == 0:
            QMessageBox.warning(
                self, 
                "Arquivo Vazio", 
                "⚠️ Não é possível gerar o arquivo!\n\n"
                "O arquivo mensal anterior está vazio.\n\n"
                "Por favor:\n"
                "• Carregue um arquivo mensal que tenha dados\n"
                "• Ou verifique se o arquivo atual tem dados válidos"
            )
            return
        
        saida = self.ed_saida.text().strip()
        if not saida:
            QMessageBox.warning(self, "Atenção", "Defina o arquivo de saída")
            return
        
        config = self._obter_config()
        if not config:
            return
        
        filtrar, areas_seco = self._obter_filtro_seco()
        if filtrar and areas_seco is None:
            return
        
        # Confirmação antes de salvar
        if len(self.df_final) == 0:
            resposta = QMessageBox.question(
                self, 
                "Confirmação", 
                "⚠️ O preview está vazio!\n\n"
                "Deseja gerar o arquivo mesmo assim?\n\n"
                "Isso criará um arquivo Excel apenas com cabeçalhos.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if resposta == QMessageBox.No:
                return
        
        self.checklist_widget.marcar_etapa("4")
        
        # Processa
        self.btn_gerar.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        try:
            self.df_final = self.processador.gerar_mensal(
                self.dados_anteriores,
                saida,
                config,
                filtrar,
                areas_seco
            )
            
            # Atualiza tabela final
            self.tabela_widget.atualizar_tabela(self.df_final)
            self.checklist_widget.marcar_etapa("5")
            
            # Muda para aba da tabela final
            self.tabs.setCurrentIndex(3)
            
            QMessageBox.information(self, "Sucesso!", 
                f"Arquivo gerado com sucesso!\n\n{saida}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", 
                f"Falha ao gerar arquivo:\n\n{str(e)}\n\n{traceback.format_exc()}")
        finally:
            self.btn_gerar.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def _log(self, mensagem: str):
        """Adiciona mensagem ao log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.txt_log.append(f"[{timestamp}] {mensagem}")
    
    def _atualizar_progresso(self, valor: int):
        """Atualiza barra de progresso"""
        self.progress_bar.setValue(valor)
        QApplication.processEvents()


# ================================== MAIN ==================================== #

def main():
    """Função principal"""
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 9))
    
    janela = MainWindow()
    janela.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
